# ─────────────────────────────────────────────────────────────────────────────
# DJANGO CORE & AUTH
# ─────────────────────────────────────────────────────────────────────────────
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date
import traceback
from django.views.decorators.http import require_POST
# ─────────────────────────────────────────────────────────────────────────────
# DJANGO HTTP & UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
from django.http import JsonResponse, HttpResponse
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.urls import reverse
# ─────────────────────────────────────────────────────────────────────────────
# DJANGO ORM & DATABASE
# ─────────────────────────────────────────────────────────────────────────────
from django.db import models, transaction
from django.db.models import Max, Sum, OuterRef, Subquery, Q
from django.db.models.functions import Coalesce, ExtractMonth, ExtractYear
from reportlab.lib import colors
# ─────────────────────────────────────────────────────────────────────────────
# PYTHON STANDARD LIBRARIES
# ─────────────────────────────────────────────────────────────────────────────
from datetime import datetime, date, timedelta
from django.utils import timezone

from decimal import Decimal
from math import ceil
from collections import defaultdict
from calendar import month_abbr, month_name
import random
import json
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors  # ✅ Corrected import
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.http import HttpResponse
import os
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
# ─────────────────────────────────────────────────────────────────────────────
# FORMS & MODELS
# ─────────────────────────────────────────────────────────────────────────────
from .forms import (
    RegisterForm,
    LoginForm,
    #ConsignmentForm,
    ConsignmentCreateForm,
    ConsignmentEditForm,
    VehicleForm,
    ConsigneeForm,
    ConsignerForm,
    LocationForm,
    GoodsInfoForm,
)
from .models import *

# ─────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
from .utils import send_otp_to_email
from num2words import num2words
# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
from openpyxl.styles import PatternFill
from reportlab.lib.colors import HexColor
# ─────────────────────────────────────────────────────────────────────────────
# DJANGO REST FRAMEWORK
# ─────────────────────────────────────────────────────────────────────────────
from rest_framework.authtoken.models import Token
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated

# ─────────────────────────────────────────────────────────────────────────────
# DRF SERIALIZERS
# ─────────────────────────────────────────────────────────────────────────────
from .serializers import RegisterSerializer, LoginSerializer, UserSerializer
from django.db.models import Sum, Q

# ─────────────────────────────────────────────────────────────────────────────
# WEB: LOGIN VIEW
# ─────────────────────────────────────────────────────────────────────────────
def login_view(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            otp = form.cleaned_data.get("otp", "")

            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                return render(request, "login.html", {"form": form, "error": "User not found"})

            if otp:
                try:
                    otp_obj = OTPModel.objects.get(email=email)
                except OTPModel.DoesNotExist:
                    return render(request, "login.html", {"form": form, "error": "OTP not found or expired"})

                if otp == otp_obj.otp:
                    login(request, user)
                    Token.objects.filter(user=user).delete()
                    Token.objects.create(user=user)
                    otp_obj.delete()
                    return redirect("dashboard")
                else:
                    return render(request, "login.html", {"form": form, "error": "Invalid OTP"})
            else:
                send_otp_to_email(email)
                return render(request, "login.html", {
                    "form": form,
                    "message": "OTP sent to your email.",
                    "otp_phase": True
                })
    else:
        form = LoginForm()
    return render(request, "login.html", {"form": form})


# ─────────────────────────────────────────────────────────────────────────────
# WEB: REGISTER VIEW
# ─────────────────────────────────────────────────────────────────────────────
def register_view(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            password = form.cleaned_data["password"]
            confirm_password = form.cleaned_data["confirm_password"]

            if password != confirm_password:
                return render(request, "register.html", {"form": form, "error": "Passwords do not match"})

            if User.objects.filter(email=email).exists():
                return render(request, "register.html", {"form": form, "error": "Email already exists"})

            user = User.objects.create_user(
                username=email.split('@')[0],
                email=email,
                password=password
            )

            RegisterModel.objects.create(
                email=email,
                password=password,
                confirm_password=confirm_password
            )

            return redirect("login")
        else:
            return render(request, "register.html", {"form": form})
    else:
        form = RegisterForm()
    return render(request, "register.html", {"form": form})


# ─────────────────────────────────────────────────────────────────────────────
# WEB: LOGOUT VIEW
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def logout_view(request):
    logout(request)
    return redirect("login")


# ─────────────────────────────────────────────────────────────────────────────
# API: VERIFY EMAIL
# ─────────────────────────────────────────────────────────────────────────────
@csrf_exempt
def verify_email_view(request):
    if request.method == "POST":
        data = json.loads(request.body)
        email = data.get("email")

        if not email:
            return JsonResponse({"status": "error", "message": "Email is required"})

        try:
            User.objects.get(email=email)
        except User.DoesNotExist:
            return JsonResponse({"status": "error", "message": "User not found"})

        otp = str(random.randint(100000, 999999))
        OTPModel.objects.update_or_create(email=email, defaults={"otp": otp})
        send_otp_to_email(email, otp)

        return JsonResponse({"status": "success", "message": "OTP sent to your email."})
    return JsonResponse({"status": "error", "message": "Invalid request method"})



# ──────────────────────────────────────────────────────────────
# CONSIGNMENT CREATE / UPDATE
# ──────────────────────────────────────────────────────────────

@login_required(login_url="login")
def consignment_form_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "message": "Invalid JSON data in request body"},
                status=400,
            )

        # Patch foreign keys into data before validation
        patched_data = data.copy()
        patched_data["vehicle_no"] = data.get("vehicle_no")
        patched_data["consignee"] = data.get("consignee")
        patched_data["consigner"] = data.get("consigner")
        patched_data["from_location"] = data.get("from_location")
        patched_data["To_Location"] = data.get("to_location")  # Match model field case

        form = ConsignmentCreateForm(patched_data)

        if form.is_valid():
            try:
                with transaction.atomic():
                    consignment_instance = form.save(commit=False)
                    consignment_instance.save()

                    # Clean previous GoodsInfo entries (for updates)
                    GoodsInfo.objects.filter(consignment=consignment_instance).delete()

                    goods_items = data.get("goods_info", [])
                    total_fare_amount = Decimal("0.00")

                    for item_data in goods_items:
                        try:
                            quantity = Decimal(str(item_data.get("quantity") or "0"))
                            rate = Decimal(str(item_data.get("rate") or "0"))
                            gi_amount = quantity * rate
                            total_fare_amount += gi_amount

                            goods_item_payload = {
                                "unit": item_data.get("unit") or "",
                                "quantity": quantity,
                                "rate": rate,
                                "gi_amount": gi_amount,
                                "from_party": item_data.get("from_party") or None,
                                "to_party": item_data.get("to_party") or None,
                            }

                            if goods_item_payload["from_party"] == "":
                                goods_item_payload["from_party"] = None
                            if goods_item_payload["to_party"] == "":
                                goods_item_payload["to_party"] = None

                            goods_form = GoodsInfoForm(goods_item_payload)
                            if goods_form.is_valid():
                                goods_info_instance = goods_form.save(commit=False)
                                goods_info_instance.consignment = consignment_instance
                                goods_info_instance.save()
                            else:
                                print(f"Invalid GoodsInfoForm for item: {goods_form.errors.as_json()}")
                        except Exception as goods_err:
                            print(f"Error saving one GoodsInfo item: {str(goods_err)}")

                    # Save total fare
                    consignment_instance.total_fare = total_fare_amount
                    consignment_instance.save(update_fields=["total_fare"])

                    # --- Vehicle existence check ---
                    submitted_vehicle_no_id = consignment_instance.Vehicle_No_id
                    vehicle_entry_date = consignment_instance.Booking_Date or timezone.now().date()

                    if submitted_vehicle_no_id:
                        try:
                            vehicle_obj_from_db = Vehicle.objects.get(pk=submitted_vehicle_no_id)
                            vehicle_number = vehicle_obj_from_db.vehicle_number
                        except Vehicle.DoesNotExist:
                            print(f"Warning: Vehicle with ID {submitted_vehicle_no_id} not found. Using ID as fallback.")
                            vehicle_number = str(submitted_vehicle_no_id)

                        try:
                            _, created = Vehicle.objects.get_or_create(
                                vehicle_number=vehicle_number,
                                defaults={"date_added": vehicle_entry_date},
                            )
                            if created:
                                print(f"Vehicle '{vehicle_number}' created.")
                            else:
                                print(f"Vehicle '{vehicle_number}' already exists.")
                        except Exception as veh_err:
                            print(f"Error saving vehicle: {str(veh_err)}")

                    # --- Consignee existence check ---
                    submitted_consignee_id = consignment_instance.consignee_id
                    if submitted_consignee_id:
                        try:
                            consignee_obj = Consignee.objects.get(pk=submitted_consignee_id)
                            consignee_name = consignee_obj.consignee_name
                        except Consignee.DoesNotExist:
                            print(f"Warning: Consignee with ID {submitted_consignee_id} not found. Using ID as fallback.")
                            consignee_name = str(submitted_consignee_id)

                        try:
                            _, created = Consignee.objects.get_or_create(
                                consignee_name=consignee_name
                            )
                            if created:
                                print(f"Consignee '{consignee_name}' created.")
                            else:
                                print(f"Consignee '{consignee_name}' already exists.")
                        except Exception as consignee_err:
                            print(f"Error ensuring consignee: {str(consignee_err)}")

                    # --- Consigner existence check ---
                    submitted_consigner_id = consignment_instance.consigner_id
                    if submitted_consigner_id:
                        try:
                            consigner_obj = Consigner.objects.get(pk=submitted_consigner_id)
                            consigner_name = consigner_obj.consigner_name
                        except Consigner.DoesNotExist:
                            print(f"Warning: Consigner with ID {submitted_consigner_id} not found. Using ID as fallback.")
                            consigner_name = str(submitted_consigner_id)

                        try:
                            _, created = Consigner.objects.get_or_create(
                                consigner_name=consigner_name
                            )
                            if created:
                                print(f"Consigner '{consigner_name}' created.")
                            else:
                                print(f"Consigner '{consigner_name}' already exists.")
                        except Exception as consigner_err:
                            print(f"Error ensuring consigner: {str(consigner_err)}")

                    # --- From Location existence check ---
                    submitted_from_location_id = consignment_instance.from_location_id
                    if submitted_from_location_id:
                        try:
                            from_location_obj = Location.objects.get(pk=submitted_from_location_id)
                            from_location_name = from_location_obj.location_name
                        except Location.DoesNotExist:
                            print(f"Warning: From Location with ID {submitted_from_location_id} not found. Using ID as fallback.")
                            from_location_name = str(submitted_from_location_id)

                        try:
                            _, created = Location.objects.get_or_create(
                                location_name=from_location_name
                            )
                            if created:
                                print(f"From Location '{from_location_name}' created.")
                            else:
                                print(f"From Location '{from_location_name}' already exists.")
                        except Exception as from_loc_err:
                            print(f"Error ensuring from location: {str(from_loc_err)}")

                    # --- To Location existence check ---
                    submitted_to_location_id = consignment_instance.To_Location_id
                    if submitted_to_location_id:
                        try:
                            to_location_obj = Location.objects.get(pk=submitted_to_location_id)
                            to_location_name = to_location_obj.location_name
                        except Location.DoesNotExist:
                            print(f"Warning: To Location with ID {submitted_to_location_id} not found. Using ID as fallback.")
                            to_location_name = str(submitted_to_location_id)

                        try:
                            _, created = Location.objects.get_or_create(
                                location_name=to_location_name
                            )
                            if created:
                                print(f"To Location '{to_location_name}' created.")
                            else:
                                print(f"To Location '{to_location_name}' already exists.")
                        except Exception as to_loc_err:
                            print(f"Error ensuring to location: {str(to_loc_err)}")

                return JsonResponse(
                    {"success": True, "message": "Consignment Form saved successfully!"},
                    status=200,
                )

            except Exception as e:
                print(f"Error during consignment save transaction: {e}")
                return JsonResponse(
                    {"success": False, "message": f"An unexpected server error occurred: {str(e)}"},
                    status=500,
                )

        else:
            print("Consignment form validation failed:", form.errors.as_json())
            return JsonResponse(
                {
                    "success": False,
                    "message": "Consignment form validation failed.",
                    "errors": form.errors.as_json(),
                },
                status=400,
            )

    else:
        form = ConsignmentCreateForm()
        context = {
            "form": form,
            "vehicles": Vehicle.objects.all(),
            "consignees": Consignee.objects.all(),
            "consigners": Consigner.objects.all(),
            "locations": Location.objects.all(),
        }
        return render(request, "consignment.html", context)
    
#_____________________________________________________________________

#----Edit Consignment form view 
#_____________________________________________________________________
@login_required(login_url="login")
def edit_consignment(request, cnid):
    instance = get_object_or_404(Consignment, CNID=cnid)

    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "message": "Invalid JSON data in request body"},
                status=400,
            )

        form = ConsignmentEditForm(data, instance=instance)

        if form.is_valid():
            form.save()
            goods_units = []
            goods_quantities = []
            goods_rates = []
            goods_from_parties = []
            goods_to_parties = []
            goods_ids = []

            for item in data.get("goods_info", []):
                goods_units.append(item.get("unit"))
                goods_quantities.append(item.get("quantity"))
                goods_rates.append(item.get("rate"))
                goods_from_parties.append(item.get("from_party"))
                goods_to_parties.append(item.get("to_party"))
                goods_ids.append(item.get("GI_ID"))

            existing_goods_ids = []

            for i in range(len(goods_units)):
                goods_id = goods_ids[i]
                try:
                    quantity = Decimal(goods_quantities[i])
                    rate = Decimal(goods_rates[i])
                    gi_amount = quantity * rate
                except:
                    quantity = Decimal("0")
                    rate = Decimal("0")
                    gi_amount = Decimal("0")

                from_party = goods_from_parties[i]
                to_party = goods_to_parties[i]

                if goods_id:
                    try:
                        item = GoodsInfo.objects.get(GI_ID=goods_id)
                        item.unit = goods_units[i]
                        item.quantity = quantity
                        item.rate = rate
                        item.gi_amount = gi_amount
                        item.from_party_id = from_party
                        item.to_party_id = to_party
                        item.save()
                        existing_goods_ids.append(item.GI_ID)
                    except GoodsInfo.DoesNotExist:
                        pass
                else:
                    item = GoodsInfo.objects.create(
                        consignment=instance,
                        unit=goods_units[i],
                        quantity=quantity,
                        rate=rate,
                        gi_amount=gi_amount,
                        from_party_id=from_party,
                        to_party_id=to_party,
                    )
                    existing_goods_ids.append(item.GI_ID)

            # Delete goods removed from frontend
            GoodsInfo.objects.filter(consignment=instance).exclude(GI_ID__in=existing_goods_ids).delete()

            return JsonResponse({"success": True})

        return JsonResponse(
            {"success": False, "message": "Form is invalid", "errors": form.errors},
            status=400,
        )

    else:
        # GET request
        form = ConsignmentEditForm(instance=instance)
        goods_items = GoodsInfo.objects.filter(consignment=instance)

        context = {
            "form": form,
            "editing_mode": True,
            "cnid": instance.CNID,
            "goods_items": goods_items,
            "vehicles": Vehicle.objects.all(),
            "consignees": Consignee.objects.all(),
            "consigners": Consigner.objects.all(),
            "locations": Location.objects.all(),
            "initial_vehicle_id": instance.Vehicle_No_id,
            "initial_consignee_id": instance.consignee_id,
            "initial_consigner_id": instance.consigner_id,
            "initial_from_location_id": instance.from_location_id,
            "initial_to_location_id": instance.To_Location_id,
        }
        return render(request, "consignment.html", context)
# ──────────────────────────────────────────────────────────────
# CONSIGNMENT DELETE
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def delete_consignment(request, cnid):
    try:
        if request.headers.get('x-requested-with') != 'XMLHttpRequest':
            return JsonResponse({"success": False, "message": "Invalid request type"}, status=400)

        consignment = get_object_or_404(Consignment, CNID=cnid)
        consignment.delete()
        return JsonResponse({"success": True})
    
    except Exception as e:
        print(" DELETE ERROR:", str(e))
        traceback.print_exc()
        return JsonResponse({"success": False, "message": "Server error."}, status=500)

# ──────────────────────────────────────────────────────────────
# CONSIGNEE CRUD
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def add_consignee(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            form = ConsigneeForm(data)
            if form.is_valid():
                consignee = form.save()
                return JsonResponse({
                    "success": True,
                    "message": "Consignee added successfully!",
                    "consignee": {
                        "id": consignee.id,
                        "consignee_name": consignee.consignee_name,
                        "date_added": consignee.date_added.strftime("%Y-%m-%d"),
                        "consignee_address": consignee.consignee_address,
                    },
                })
            return JsonResponse({"success": False, "errors": form.errors.as_json()}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)
    return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)


@login_required(login_url="login")
def get_consignees(request):
    if request.method == "GET":
        consignees = Consignee.objects.all().values("id", "consignee_name", "date_added", "consignee_address")
        return JsonResponse({
            "success": True,
            "consignees": [
                {
                    "id": c["id"],
                    "consignee_name": c["consignee_name"],
                    "date_added": c["date_added"].strftime("%Y-%m-%d"),
                    "consignee_address": c["consignee_address"],
                } for c in consignees
            ]
        })
    return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)


@login_required(login_url="login")
@require_http_methods(["DELETE"])
def delete_consignee(request, consignee_id):
    try:
        consignee = Consignee.objects.get(id=consignee_id)
        consignee.delete()
        return JsonResponse({"success": True, "message": "Consignee deleted successfully!"})
    except Consignee.DoesNotExist:
        return JsonResponse({"success": False, "message": "Consignee not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)


# ──────────────────────────────────────────────────────────────
# CONSIGNER CRUD
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def add_consigner(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            form = ConsignerForm(data)
            if form.is_valid():
                consigner = form.save()
                return JsonResponse({
                    "success": True,
                    "message": "Consigner added successfully!",
                    "consigner": {
                        "id": consigner.id,
                        "consigner_name": consigner.consigner_name,
                        "date_added": consigner.date_added.strftime("%Y-%m-%d"),
                        "consigner_address": consigner.consigner_address,
                    },
                })
            return JsonResponse({"success": False, "errors": form.errors.as_json()}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)
    return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)


@login_required(login_url="login")
def get_consigners(request):
    if request.method == "GET":
        consigners = Consigner.objects.all().values("id", "consigner_name", "date_added", "consigner_address")
        return JsonResponse({
            "success": True,
            "consigners": [
                {
                    "id": c["id"],
                    "consigner_name": c["consigner_name"],
                    "date_added": c["date_added"].strftime("%Y-%m-%d"),
                    "consigner_address": c["consigner_address"],
                } for c in consigners
            ]
        })
    return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)


@login_required(login_url="login")
@require_http_methods(["DELETE"])
def delete_consigner(request, consigner_id):
    try:
        consigner = Consigner.objects.get(id=consigner_id)
        consigner.delete()
        return JsonResponse({"success": True, "message": "Consigner deleted successfully!"})
    except Consigner.DoesNotExist:
        return JsonResponse({"success": False, "message": "Consigner not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

# ──────────────────────────────────────────────────────────────
# LOCATION MANAGEMENT
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def location_list_view(request):
    # Fetch all locations directly from the Location model
    all_locations = Location.objects.all().order_by("location_name")

    # Prepare data for template
    locations_for_template = []
    for loc_obj in all_locations:
        locations_for_template.append(
            {
                "id": loc_obj.id,
                "location_name": loc_obj.location_name,
                "date_added": (
                    loc_obj.date_added.strftime("%Y-%m-%d")
                    if loc_obj.date_added
                    else ""
                ),  # Format date
                "type": "Location",  # General type, as it's not specific to 'from' or 'to'
            }
        )

    # Pass a single list of all locations to the template
    return render(
        request,
        "location_list.html",
        {
            "locations": locations_for_template,  # Changed context variable name
        },
    )


def edit_location(request):
    location_id = request.POST.get("location_id")
    name = request.POST.get("location_name")
    date_added = request.POST.get("date_added")
    try:
        loc = Location.objects.get(id=location_id)
        loc.location_name = name
        loc.date_added = date_added
        loc.save()
        return JsonResponse({"success": True})
    except Location.DoesNotExist:
        return JsonResponse({"success": False, "error": "Location not found"})

@login_required(login_url="login")
@require_http_methods(["POST"])
def add_location(request):
    try:
        data = json.loads(request.body)
        form = LocationForm(data)
        if form.is_valid():
            location = form.save()
            return JsonResponse({
                "success": True,
                "message": "Location added successfully!",
                "location": {
                    "id": location.id,
                    "location_name": location.location_name,
                    "date_added": location.date_added.strftime("%Y-%m-%d"),
                },
            })
        return JsonResponse({"success": False, "errors": form.errors.as_json()}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)


@login_required(login_url="login")
@require_http_methods(["GET"])
def get_locations(request):
    locations = Location.objects.all().values("id", "location_name", "date_added")
    return JsonResponse({
        "success": True,
        "locations": [
            {
                "id": loc["id"],
                "location_name": loc["location_name"],
                "date_added": loc["date_added"].strftime("%Y-%m-%d"),
            } for loc in locations
        ]
    })


@login_required(login_url="login")
@require_http_methods(["DELETE"])
def delete_location(request, location_id):
    try:
        location = Location.objects.get(id=location_id)
        location.delete()
        return JsonResponse({"success": True, "message": "Location deleted successfully!"})
    except Location.DoesNotExist:
        return JsonResponse({"success": False, "error": "Location not found"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
    
    
# ──────────────────────────────────────────────────────────────
# DASHBOARD VIEW
# ──────────────────────────────────────────────────────────────

@login_required(login_url="login")
def dashboard_view(request):
    try:
        selected_year = int(request.GET.get("year", date.today().year))
    except (ValueError, TypeError):
        selected_year = datetime.date.today().year

    goods_filtered = GoodsInfo.objects.filter(consignment__Booking_Date__year=selected_year)
    goods_all = GoodsInfo.objects.filter(consignment__Booking_Date__year__in=[2024, 2025])

    years = Consignment.objects.annotate(year=ExtractYear("Booking_Date")) \
        .filter(year__gte=2024).values_list("year", flat=True).distinct().order_by("-year")

    summary = goods_all.aggregate(
        total_advance=Sum("paid_amount"),
        total_fare=Sum("gi_amount"),
    )
    summary["total_advance"] = summary["total_advance"] or 0
    summary["total_fare"] = summary["total_fare"] or 0
    summary["total_balance"] = summary["total_fare"] - summary["total_advance"]

    # Monthly Chart
    monthly_data = goods_filtered.annotate(month=ExtractMonth("consignment__Booking_Date")) \
        .values("month").annotate(
            total_advance=Sum("paid_amount"),
            total_fare=Sum("gi_amount"),
        ).order_by("month")

    month_labels, advance_values, balance_values, fare_values = [], [], [], []
    monthly_summary = {entry["month"]: entry for entry in monthly_data}

    for i in range(1, 13):
        month_labels.append(month_abbr[i])
        entry = monthly_summary.get(i)
        advance = float(entry["total_advance"] or 0) if entry else 0
        fare = float(entry["total_fare"] or 0) if entry else 0
        balance = fare - advance
        advance_values.append(advance)
        balance_values.append(balance)
        fare_values.append(fare)

    # Truck Summary (Inam added here ✅)
    truck_totals = Consignment.objects.filter(Booking_Date__year__in=[2024, 2025]) \
        .values("Vehicle_No__vehicle_number") \
        .annotate(
            total_advance=Sum("Advance_Amount"),
            total_balance=Sum("Balance_Amount"),
            total_fare=Sum("Truck_Freight"),
            total_innam=Sum("Innam")  # ✅ Added line
        ).order_by("Vehicle_No__vehicle_number")

    for item in truck_totals:
        item["vehicle_number"] = item.pop("Vehicle_No__vehicle_number")

    total_truck_advance = sum(t["total_advance"] or 0 for t in truck_totals)
    total_truck_balance = sum(t["total_balance"] or 0 for t in truck_totals)
    total_truck_fare = sum(t["total_fare"] or 0 for t in truck_totals)

    # Party Summary (From → To)
    grouped_data = {}
    goods_queryset = GoodsInfo.objects.select_related("from_party", "to_party") \
        .filter(consignment__Booking_Date__year__in=[2024, 2025], to_party__isnull=False)

    for item in goods_queryset:
        key = (item.from_party, item.to_party)
        grouped_data.setdefault(key, {"amount": 0, "paid": 0})
        grouped_data[key]["amount"] += item.gi_amount or 0
        grouped_data[key]["paid"] += item.paid_amount or 0

    sorted_grouped = sorted(
        grouped_data.items(),
        key=lambda x: ((x[0][1].consignee_name or "") if x[0][1] else "").lower()
    )

    goods_summary_grouped = []
    for i, ((from_party, to_party), totals) in enumerate(sorted_grouped, 1):
        amount = totals["amount"]
        paid = totals["paid"]
        due = amount - paid
        goods_summary_grouped.append({
            "sr": i,
            "from_party": from_party,
            "to_party": to_party,
            "total_gi": amount,
            "total_paid": paid,
            "total_balance": due,
        })

    total_gi = sum(g["total_gi"] for g in goods_summary_grouped)
    total_paid = sum(g["total_paid"] for g in goods_summary_grouped)

    context = {
        "years": list(years),
        "selected_year": selected_year,
        "summary": summary,
        "truck_totals": truck_totals,
        "truck_totals_summary": {
            "advance": total_truck_advance,
            "balance": total_truck_balance,
            "fare": total_truck_fare,
        },
        "month_labels": json.dumps(month_labels),
        "advance_values": json.dumps(advance_values),
        "balance_values": json.dumps(balance_values),
        "fare_values": json.dumps(fare_values),
        "goods_summary_grouped": goods_summary_grouped,
        "goods_summary_total": {
            "gi_amount": total_gi,
            "paid_amount": total_paid,
            "balance_amount": total_gi - total_paid,
        }
    }

    return render(request, "dashboard.html", context)
# ──────────────────────────────────────────────────────────────
#  Party-wise Summary PDF View
# ─────────────────────────────────────────────────────────────
@login_required(login_url="login")
def partywise_summary_pdf(request, year):
    # ─── Get Query Parameter ──────────────────────────
    party_query = request.GET.get("party", "").strip().lower()

    # ─── PDF Setup ─────────────────────────────────────
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f"attachment; filename=PartyWiseSummary_{year}.pdf"

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1 * cm, rightMargin=1 * cm,
        topMargin=1.5 * cm, bottomMargin=1 * cm
    )

    styles = getSampleStyleSheet()
    elements = []

    # ─── Header ──────────────────────────────────────
    company_style = styles['Title'].clone('CompanyStyle')
    company_style.fontSize = 16
    company_style.textColor = colors.HexColor('#1a365d')
    company_style.alignment = TA_CENTER

    subtitle_style = styles['Normal'].clone('SubtitleStyle')
    subtitle_style.fontSize = 10
    subtitle_style.alignment = TA_CENTER

    address_style = styles['Normal'].clone('AddressStyle')
    address_style.fontSize = 9
    address_style.alignment = TA_CENTER

    logo_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
    ]

    logo_cell = None
    for logo_path in logo_paths:
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3 * cm, height=2.5 * cm)
            logo.hAlign = 'RIGHT'
            logo_cell = logo
            break
    else:
        logo_cell = Paragraph("DHANLAXMI", company_style)

    address_block = [
        Paragraph("<b>DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS</b>", company_style),
        Paragraph("<b>IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING</b>", subtitle_style),
        Paragraph("18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304", address_style),
        Paragraph("Mob: +91 9830357491 | Email: kakashri7314@gmail.com", address_style),
    ]

    header_table = Table([[address_block, logo_cell]], colWidths=[16 * cm, 5 * cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))

    # ─── Title ────────────────────────────────────────
    title_style = styles['Title'].clone('TitleStyle')
    title_style.fontSize = 14
    title_style.textColor = colors.HexColor('#6a1b9a')
    title_style.alignment = TA_CENTER

    # ✅ Always display plain title (no filtering text)
    title_text = f"<b>Party-wise Summary - {year}</b>"
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 12))

    # ─── Data Aggregation ─────────────────────────────
    goods_queryset = GoodsInfo.objects.select_related("from_party", "to_party") \
        .filter(consignment__Booking_Date__year=year, to_party__isnull=False)

    # Apply filter if party is searched
    if party_query:
        goods_queryset = goods_queryset.filter(
            Q(from_party__consignee_name__icontains=party_query) |
            Q(to_party__consignee_name__icontains=party_query)
        )

    grouped_data = {}
    for item in goods_queryset:
        key = (item.from_party, item.to_party)
        grouped_data.setdefault(key, {"amount": 0, "paid": 0})
        grouped_data[key]["amount"] += item.gi_amount or 0
        grouped_data[key]["paid"] += item.paid_amount or 0

    # Table Header
    data = [["Sr No", "From Party", "To Party", "Amount", "Paid", "Balance"]]

    # Totals
    total_amount = 0
    total_paid = 0
    total_balance = 0

    for i, ((from_party, to_party), totals) in enumerate(grouped_data.items(), 1):
        amount = totals["amount"]
        paid = totals["paid"]
        balance = amount - paid
        total_amount += amount
        total_paid += paid
        total_balance += balance

        data.append([
            i,
            from_party.consignee_name if from_party else "—",
            to_party.consignee_name if to_party else "—",
            f"{amount:,.2f}",
            f"{paid:,.2f}",
            f"{balance:,.2f}"
        ])

    # Totals Row
    data.append([
        "", "", "Total",
        f"{total_amount:,.2f}",
        f"{total_paid:,.2f}",
        f"{total_balance:,.2f}"
    ])

    # ─── Table Rendering ──────────────────────────────
    table = Table(data, repeatRows=1, colWidths=[2.5*cm, 6*cm, 6*cm, 4*cm, 4*cm, 4*cm])
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (-6, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    # ─── Footer ───────────────────────────────────────
    footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    footer_style = styles['Normal'].clone('FooterStyle')
    footer_style.fontSize = 8
    footer_style.textColor = colors.grey
    footer_style.alignment = TA_CENTER
    elements.append(Paragraph(footer_text, footer_style))

    # ─── Build PDF ────────────────────────────────────
    doc.build(elements)
    return response

# ─────────────────────────────────────────────────────────────
# Party-wise Summary Excel View
# ────────────────────────────────────────────────────────────
@login_required(login_url="login")
def partywise_summary_excel(request, year):
    # --- Initialize Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"Party Summary {year}"

    # --- Header: Company Info ---
    ws.merge_cells('A1:F1')
    ws.merge_cells('A2:F2')
    ws.merge_cells('A3:F3')
    ws.merge_cells('A4:F4')

    ws['A1'] = "DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS"
    ws['A2'] = "IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING"
    ws['A3'] = "18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304"
    ws['A4'] = "Mob: +91 9830357491 | Email: kakashri7314@gmail.com"

    for i in range(1, 5):
        ws[f"A{i}"].font = Font(bold=(i == 1), size=11)
        ws[f"A{i}"].alignment = Alignment(horizontal='center')

    # --- Add Logo if available ---
    try:
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
        ]
        for path in logo_paths:
            if os.path.exists(path):
                img = XLImage(path)
                img.width = 100
                img.height = 75
                ws.add_image(img, 'F1')
                break
    except Exception:
        pass

    ws.append([])  # Spacer Row

    # --- Title ---
    ws.merge_cells('A6:F6')
    ws['A6'] = f"Party-wise Summary Report - {year}"
    ws['A6'].font = Font(bold=True, size=14, color="0D47A1")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.append([])

    # --- Table Headers ---
    headers = ["Sr. No", "From Party", "To Party", "Amount (₹)", "Paid (₹)", "Balance (₹)"]
    ws.append(headers)

    header_fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # --- Filter & Prepare Data ---
    party_query = request.GET.get("party", "").strip().lower()
    goods_filtered = GoodsInfo.objects.filter(consignment__Booking_Date__year=year)

    if party_query:
        goods_filtered = goods_filtered.filter(to_party__consignee_name__icontains=party_query)

    summary_data = defaultdict(lambda: {"gi": 0, "paid": 0, "balance": 0})

    for gi in goods_filtered.select_related("from_party", "to_party"):
        key = (
            gi.from_party.consignee_name if gi.from_party else "",
            gi.to_party.consignee_name if gi.to_party else "",
        )

        gi_amt = gi.gi_amount or 0
        paid_amt = gi.paid_amount or 0
        balance_amt = gi_amt - paid_amt

        summary_data[key]["gi"] += gi_amt
        summary_data[key]["paid"] += paid_amt
        summary_data[key]["balance"] += balance_amt

    # --- Populate Rows ---
    for idx, ((from_party, to_party), data) in enumerate(summary_data.items(), 1):
        row = [
            idx,
            from_party,
            to_party,
            round(data["gi"], 2),
            round(data["paid"], 2),
            round(data["balance"], 2),
        ]
        ws.append(row)

    # --- Grand Total Row ---
    ws.append([])  # Spacer Row
    total_gi = sum(d["gi"] for d in summary_data.values())
    total_paid = sum(d["paid"] for d in summary_data.values())
    total_balance = sum(d["balance"] for d in summary_data.values())

    total_row = ["", "", "Grand Total", round(total_gi, 2), round(total_paid, 2), round(total_balance, 2)]
    ws.append(total_row)
    for i, cell in enumerate(ws[ws.max_row], start=1):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right' if i >= 4 else 'center')

    # --- Auto-fit Columns ---
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 25)

    # --- Return Excel File ---
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"PartyWiseSummary_{year}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response
# ──────────────────────────────────────────────────────────────
# NEXT CNID FETCH
# ──────────────────────────────────────────────────────────────

@login_required(login_url="login")
def get_next_cnid(request):
    if request.method == "GET":
        try:
            last_cn = Consignment.objects.order_by('-CNID').first()
            next_cn_id = (last_cn.CNID + 1) if last_cn else 1
            return JsonResponse({"success": True, "next_cn_id": next_cn_id})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)
    else:
        return JsonResponse({"success": False, "message": "Invalid request method"}, status=405)

# ──────────────────────────────────────────────────────────────
# VEHICLE MANAGEMENT VIEWS
# ──────────────────────────────────────────────────────────────

def vehicle_list(request):
    """Render the vehicle list page."""
    return render(request, "vehicle_list.html")


@require_http_methods(["GET"])
def get_vehicles(request):
    """
    Return all vehicles as JSON.
    Useful for dynamic frontend rendering via JavaScript.
    """
    vehicles = Vehicle.objects.all().order_by("-date_added")
    formatted = [
        {
            "id": v.id,
            "vehicle_number": v.vehicle_number,
            "owner_name": v.owner_name,
            "owner_phone": v.owner_phone,
            "date_added": v.date_added.strftime("%Y-%m-%d"),
        }
        for v in vehicles
    ]
    return JsonResponse({"success": True, "vehicles": formatted})


@csrf_exempt
@require_http_methods(["POST"])
def add_vehicle(request):
    """
    Add a new vehicle.
    Optional: owner_name, owner_phone
    Required: vehicle_number, date_added
    """
    try:
        data = json.loads(request.body)
        vehicle_number = data.get("vehicle_number")
        date_added = parse_date(data.get("date_added"))
        owner_name = data.get("owner_name") or None
        owner_phone = data.get("owner_phone") or None

        if not vehicle_number or not date_added:
            return JsonResponse({"success": False, "message": "Missing required fields."})

        vehicle = Vehicle.objects.create(
            vehicle_number=vehicle_number,
            owner_name=owner_name,
            owner_phone=owner_phone,
            date_added=date_added,
        )

        return JsonResponse({
            "success": True,
            "vehicle": {
                "id": vehicle.id,
                "vehicle_number": vehicle.vehicle_number,
                "owner_name": vehicle.owner_name,
                "owner_phone": vehicle.owner_phone,
                "date_added": vehicle.date_added.strftime("%Y-%m-%d"),
            }
        })

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def edit_vehicle(request, vehicle_id):
    """
    Edit an existing vehicle.
    Requires: vehicle_number, date_added
    Optional: owner_name, owner_phone
    """
    try:
        data = json.loads(request.body)
        vehicle_number = data.get("vehicle_number")
        date_added = parse_date(data.get("date_added"))
        owner_name = data.get("owner_name") or None
        owner_phone = data.get("owner_phone") or None

        if not vehicle_number or not date_added:
            return JsonResponse({"success": False, "message": "Missing required fields."})

        vehicle = get_object_or_404(Vehicle, id=vehicle_id)
        vehicle.vehicle_number = vehicle_number
        vehicle.date_added = date_added
        vehicle.owner_name = owner_name
        vehicle.owner_phone = owner_phone
        vehicle.save()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def delete_vehicle(request, vehicle_id):
    """
    Delete a vehicle by ID using POST request.
    """
    try:
        vehicle = get_object_or_404(Vehicle, id=vehicle_id)
        vehicle.delete()
        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


# ──────────────────────────────────────────────────────────────
# party VIEW
# ──────────────────────────────────────────────────────────────
from itertools import chain
from operator import attrgetter

def party_list_view(request):
    consignees = list(Consignee.objects.all())
    consigners = list(Consigner.objects.all())

    # Tag party type and unify fields
    for c in consignees:
        c.party_type = 'consignee'
        c.party_name = c.consignee_name
        c.party_address = c.consignee_address

    for c in consigners:
        c.party_type = 'consigner'
        c.party_name = c.consigner_name
        c.party_address = c.consigner_address
        c.contact_no = None

    combined_parties = sorted(
        chain(consignees, consigners),
        key=attrgetter("date_added"),
        reverse=True,
    )

    paginator = Paginator(combined_parties, 25)
    page_number = int(request.GET.get("page", 1))
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages
    max_pages_display = 4  # Show 4 pages at a time

    start_page = max(1, page_number - (max_pages_display // 2))
    end_page = min(total_pages, start_page + max_pages_display - 1)

    # Adjust start_page if we're near the end
    if end_page - start_page < max_pages_display - 1:
        start_page = max(1, end_page - max_pages_display + 1)

    page_range = range(start_page, end_page + 1)

    return render(
        request,
        "party_list.html",
        {
            "page_obj": page_obj,
            "current_page": page_number,
            "page_range": page_range,
            "start_page": start_page,
            "end_page": end_page,
            "total_pages": total_pages,
        },
    )

#____________________________________________________________________

def edit_party(request):
    party_id = request.POST.get("party_id")
    party_type = request.POST.get("party_type")
    name = request.POST.get("party_name")
    address = request.POST.get("party_address")
    date_added = request.POST.get("date_added")
    contact_no = request.POST.get("contact_no")  #  new line to accept contact no.

    model = Consignee if party_type == "consignee" else Consigner

    try:
        instance = model.objects.get(id=party_id)
        if party_type == "consignee":
            instance.consignee_name = name
            instance.consignee_address = address
            instance.contact_no = contact_no  #  update contact no.
        else:
            instance.consigner_name = name
            instance.consigner_address = address
            # Optional: handle contact_no if added to Consigner model in future

        instance.date_added = date_added
        instance.save()
        return JsonResponse({"success": True})
    except model.DoesNotExist:
        return JsonResponse({"success": False, "error": "Party not found"})

def delete_party(request, party_type, party_id):
    if request.method == "DELETE":
        try:
            if party_type == "consignee":
                Consignee.objects.get(id=party_id).delete()
            elif party_type == "consigner":
                Consigner.objects.get(id=party_id).delete()
            else:
                return JsonResponse({"success": False, "error": "Invalid party type"})
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request"})

# ──────────────────────────────────────────────────────────────
# CN_DATA VIEW
# ──────────────────────────────────────────────────────────────

@login_required(login_url="login")
def cn_data_view(request):
    # ✅ 1. Get global search input
    search_query = request.GET.get("global_search", "").strip()

    # ✅ 2. Base queryset
    consignment_list_raw = Consignment.objects.all().order_by("-Booking_Date")

    # ✅ 3. Apply vehicle number filter
    if search_query:
        matching_vehicle_ids = Vehicle.objects.filter(
            vehicle_number__icontains=search_query
        ).values_list("id", flat=True)
        consignment_list_raw = (
            consignment_list_raw.filter(Vehicle_No__in=matching_vehicle_ids)
            if matching_vehicle_ids else consignment_list_raw.none()
        )

    # ✅ 4. Lookup tables
    vehicles_lookup = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees_lookup = {str(c.id): c.consignee_name for c in Consignee.objects.all()}
    consigners_lookup = {str(cs.id): cs.consigner_name for cs in Consigner.objects.all()}
    locations_lookup = {str(l.id): l.location_name for l in Location.objects.all()}

    # ✅ 5. Format data
    formatted_consignments = []
    for c in consignment_list_raw:
        formatted_consignments.append({
            "CNID": c.CNID,
            "Cn_No": c.Cn_No,
            "Vehicle_No": vehicles_lookup.get(str(c.Vehicle_No), c.Vehicle_No),
            "consignee": consignees_lookup.get(str(c.consignee), c.consignee),
            "consigner": consigners_lookup.get(str(c.consigner), c.consigner),
            "from_location": locations_lookup.get(str(c.from_location), c.from_location),
            "To_Location": locations_lookup.get(str(c.To_Location), c.To_Location),
            "Booking_Date": c.Booking_Date,
            "Loading_Date": c.Loading_Date,
            "Unloading_Date": c.Unloading_Date,
            "Truck_Freight": c.Truck_Freight,
            "Advance_Amount": c.Advance_Amount,
            "Balance_Amount": c.Balance_Amount,
        })

    # ✅ 6. Pagination
    paginator = Paginator(formatted_consignments, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    current_page = page_obj.number
    total_pages = paginator.num_pages
    start_page = max(current_page - 2, 1)
    end_page = min(current_page + 2, total_pages)
    page_range = range(start_page, end_page + 1)

    query_params = request.GET.copy()
    query_params.pop("page", None)

    # ✅ 7. Pass context to template
    return render(request, "CN_Data.html", {
        "page_obj": page_obj,
        "consignments": page_obj.object_list,
        "current_page": current_page,
        "total_pages": total_pages,
        "start_page": start_page,
        "end_page": end_page,
        "page_range": page_range,
        "global_search": search_query,
        "query_string": query_params.urlencode(),
    })

# ──────────────────────────────────────────────────────────────
# export_consignments_excel VIEW
# ──────────────────────────────────────────────────────────────

@login_required(login_url="login")
def export_consignments_excel(request):
    # Lookup dictionaries
    vehicles_lookup = {str(v.id): str(v.vehicle_number) for v in Vehicle.objects.all()}
    consignees_lookup = {str(c.id): str(c.consignee_name) for c in Consignee.objects.all()}
    consigners_lookup = {str(cs.id): str(cs.consigner_name) for cs in Consigner.objects.all()}
    locations_lookup = {str(l.id): str(l.location_name) for l in Location.objects.all()}

    consignments = Consignment.objects.all().order_by("-Booking_Date")

    wb = Workbook()
    ws = wb.active
    ws.title = "Consignments"

    # --- Header: Company Info ---
    ws.merge_cells('A1:M1')
    ws.merge_cells('A2:M2')
    ws.merge_cells('A3:M3')
    ws.merge_cells('A4:M4')

    ws['A1'] = "DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS"
    ws['A2'] = "IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING"
    ws['A3'] = "18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304"
    ws['A4'] = "Mob: +91 9830357491 | Email: kakashri7314@gmail.com"

    for i in range(1, 5):
        ws[f"A{i}"].font = Font(bold=(i == 1), size=11)
        ws[f"A{i}"].alignment = Alignment(horizontal='center')

    # --- Logo ---
    try:
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
        ]
        for path in logo_paths:
            if os.path.exists(path):
                img = XLImage(path)
                img.width = 100
                img.height = 75
                ws.add_image(img, 'M1')
                break
    except Exception:
        pass

    ws.append([])  # spacer

    # --- Title Row ---
    ws.merge_cells('A6:M6')
    ws['A6'] = "Consignment Report"
    ws['A6'].font = Font(bold=True, size=14, color="0D47A1")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.append([])

    # --- Table Headers ---
    headers = [
        "S.No",
        "CN No",
        "Vehicle No",
        "Consignee",
        "From Location",
        "To Location",
        "Loading Date",
        "Unloading Date",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # --- Table Data ---
    for idx, c in enumerate(consignments, start=1):
        row = [
            idx,
            str(c.Cn_No),
            vehicles_lookup.get(str(c.Vehicle_No), str(c.Vehicle_No)),
            consignees_lookup.get(str(c.consignee), str(c.consignee)),
            locations_lookup.get(str(c.from_location), str(c.from_location)),
            locations_lookup.get(str(c.To_Location), str(c.To_Location)),
            str(c.Loading_Date) if c.Loading_Date else "",
            str(c.Unloading_Date) if c.Unloading_Date else "",
        ]
        ws.append(row)

    # --- Auto-fit Columns ---
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)

    # --- Return Response ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=consignment_data.xlsx"
    wb.save(response)
    return response

# ──────────────────────────────────────────────────────────────
# export_consignments_pdf VIEW
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def export_consignments_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="consignment_data.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1 * cm
    )
    elements = []
    styles = getSampleStyleSheet()

    # Header Styles
    company_style = styles['Title'].clone('CompanyStyle')
    company_style.fontSize = 16
    company_style.textColor = colors.HexColor('#1a365d')
    company_style.alignment = TA_CENTER

    subtitle_style = styles['Normal'].clone('SubtitleStyle')
    subtitle_style.fontSize = 10
    subtitle_style.alignment = TA_CENTER

    address_style = styles['Normal'].clone('AddressStyle')
    address_style.fontSize = 9
    address_style.alignment = TA_CENTER

    # Logo
    logo_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
    ]
    logo_cell = None
    for logo_path in logo_paths:
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3 * cm, height=2.5 * cm)
            logo.hAlign = 'RIGHT'
            logo_cell = logo
            break
    else:
        logo_cell = Paragraph("DHANLAXMI", company_style)

    # Address Block
    address_block = [
        Paragraph("<b>DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS</b>", company_style),
        Paragraph("<b>IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING</b>", subtitle_style),
        Paragraph("18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304", address_style),
        Paragraph("Mob: +91 9830357491 | Email: kakashri7314@gmail.com", address_style),
    ]

    # Header Table
    header_table = Table([[address_block, logo_cell]], colWidths=[16 * cm, 5 * cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))

    # Title
    title_style = styles['Title'].clone('TitleStyle')
    title_style.fontSize = 14
    title_style.textColor = colors.HexColor('#6a1b9a')
    title_style.alignment = TA_CENTER
    elements.append(Paragraph("<b>Consignment Report</b>", title_style))
    elements.append(Spacer(1, 12))

    # Table Headers
    headers = [
        "S.No", "CN No", "Vehicle No", "Consignee",
        "From Location", "To Location",
        "Loading Date", "Unloading Date"
    ]
    data = [headers]

    # Lookups
    vehicles = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees = {str(c.id): c.consignee_name for c in Consignee.objects.all()}
    locations = {str(l.id): l.location_name for l in Location.objects.all()}

    # Data Rows
    consignments = Consignment.objects.all().order_by("-Booking_Date")
    for i, c in enumerate(consignments, 1):
        data.append([
            i,
            c.Cn_No,
            vehicles.get(str(c.Vehicle_No), c.Vehicle_No),
            consignees.get(str(c.consignee), c.consignee),
            locations.get(str(c.from_location), c.from_location),
            locations.get(str(c.To_Location), c.To_Location),
            str(c.Loading_Date) if c.Loading_Date else "",
            str(c.Unloading_Date) if c.Unloading_Date else "",
        ])

    # Wider column widths for a bigger table
    col_widths = [2.0 * cm, 3.0 * cm, 3.0 * cm, 4.0 * cm, 4.0 * cm, 4.0 * cm, 3.0 * cm, 3.0 * cm]

    # Table Style
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Removed background
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Bold header
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),      # Normal rest
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))

    # Footer Timestamp
    footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    footer_style = styles['Normal'].clone('FooterStyle')
    footer_style.fontSize = 8
    footer_style.textColor = colors.grey
    footer_style.alignment = TA_CENTER
    elements.append(Paragraph(footer_text, footer_style))

    doc.build(elements)
    return response


# ──────────────────────────────────────────────────────────────
# REPORTS VIEW
# ──────────────────────────────────────────────────────────────


# @login_required(login_url="login")
# def reports_data_view(request):
#     consignment_list = Consignment.objects.prefetch_related('goods_items').order_by("CNID")
#     global_search_query = request.GET.get("global_search", "").strip()

#     if global_search_query:
#         q_objects = Q(Cn_No__icontains=global_search_query)

#         matching_vehicle_ids = Vehicle.objects.filter(
#             vehicle_number__icontains=global_search_query
#         ).values_list("id", flat=True)
#         if matching_vehicle_ids:
#             q_objects |= Q(Vehicle_No__in=matching_vehicle_ids)

#         matching_consignee_ids = Consignee.objects.filter(
#             consignee_name__icontains=global_search_query
#         ).values_list("id", flat=True)
#         if matching_consignee_ids:
#             q_objects |= Q(consignee__in=matching_consignee_ids)

#         matching_consigner_ids = Consigner.objects.filter(
#             consigner_name__icontains=global_search_query
#         ).values_list("id", flat=True)
#         if matching_consigner_ids:
#             q_objects |= Q(consigner__in=matching_consigner_ids)

#         matching_location_ids = Location.objects.filter(
#             location_name__icontains=global_search_query
#         ).values_list("id", flat=True)
#         if matching_location_ids:
#             q_objects |= Q(from_location__in=matching_location_ids) | Q(To_Location__in=matching_location_ids)

#         consignment_list = consignment_list.filter(q_objects)

#     # Lookup dictionaries
#     vehicles_lookup = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
#     consignees_lookup = {str(c.id): c.consignee_name for c in Consignee.objects.all()}
#     consigners_lookup = {str(cs.id): cs.consigner_name for cs in Consigner.objects.all()}
#     locations_lookup = {str(l.id): l.location_name for l in Location.objects.all()}

#     formatted_consignments = []

#     for c in consignment_list:
#         goods_infos = c.goods_items.all()

#         # If there are goods entries, create a row per goods_info
#         if goods_infos.exists():
#             for gi in goods_infos:
#                 formatted_consignments.append({
#                     "CNID": c.CNID,
#                     "Cn_No": c.Cn_No,
#                     "Vehicle_No": vehicles_lookup.get(str(c.Vehicle_No.id), str(c.Vehicle_No)),
#                     "consignee": consignees_lookup.get(str(c.consignee.id), str(c.consignee)),
#                     "consigner": consigners_lookup.get(str(c.consigner.id), str(c.consigner)),
#                     "from_location": locations_lookup.get(str(c.from_location_id), str(c.from_location_id)),
#                     "To_Location": locations_lookup.get(str(c.To_Location_id), str(c.To_Location_id)),
#                     "Booking_Date": c.Booking_Date.strftime("%Y-%m-%d") if c.Booking_Date else "",
#                     "Loading_Date": c.Loading_Date.strftime("%Y-%m-%d") if c.Loading_Date else "",
#                     "Unloading_Date": c.Unloading_Date.strftime("%Y-%m-%d") if c.Unloading_Date else "",
#                     "Truck_Freight": c.Truck_Freight,
#                     "Advance_Amount": c.Advance_Amount,
#                     "Balance_Amount": c.Balance_Amount,
#                     "from_party": gi.from_party.consignee_name if gi.from_party else "",
#                     "to_party": gi.to_party.consignee_name if gi.to_party else "",
#                 })
                

#         else:
#             # If no goods_info, still add one row with blank parties
#             formatted_consignments.append({
#                 "CNID": c.CNID,
#                 "Cn_No": c.Cn_No,
#                 "Vehicle_No": vehicles_lookup.get(str(c.Vehicle_No.id), str(c.Vehicle_No)),
#                 "consignee": consignees_lookup.get(str(c.consignee.id), str(c.consignee)),
#                 "consigner": consigners_lookup.get(str(c.consigner.id), str(c.consigner)),
#                 "from_location": locations_lookup.get(c.from_location, c.from_location),
#                 "To_Location": locations_lookup.get(c.To_Location, c.To_Location),
#                 "Booking_Date": c.Booking_Date.strftime("%Y-%m-%d") if c.Booking_Date else "",
#                 "Loading_Date": c.Loading_Date.strftime("%Y-%m-%d") if c.Loading_Date else "",
#                 "Unloading_Date": c.Unloading_Date.strftime("%Y-%m-%d") if c.Unloading_Date else "",
#                 "Truck_Freight": c.Truck_Freight,
#                 "Advance_Amount": c.Advance_Amount,
#                 "Balance_Amount": c.Balance_Amount,
#                 "from_party": "",
#                 "to_party": "",
#             })

#     # Pagination logic
#     paginator = Paginator(formatted_consignments, 10)
#     page_number = int(request.GET.get("page", 1))
#     page_obj = paginator.get_page(page_number)

#     total_pages = paginator.num_pages
#     window_size = 4
#     current_window = ceil(page_number / window_size)
#     start_page = (current_window - 1) * window_size + 1
#     end_page = min(start_page + window_size - 1, total_pages)
#     custom_page_range = range(start_page, end_page + 1)

#     query_params = request.GET.copy()
#     if "page" in query_params:
#         del query_params["page"]

#     current_filters = {
#         "global_search": global_search_query,
#     }

#     return render(request, "reports.html", {
#         "page_obj": page_obj,
#         "query_string": query_params.urlencode() if query_params else "",
#         "current_filters": current_filters,
#         "page_range": custom_page_range,
#         "current_page": page_number,
#         "start_page": start_page,
#         "end_page": end_page,
#         "total_pages": total_pages,
#     })


@login_required(login_url="login")
def reports_data_view(request):
    consignment_list = Consignment.objects.prefetch_related("goods_items").order_by("CNID")

    global_search_query = request.GET.get("global_search", "").strip()

    if global_search_query:
        q_objects = Q(Cn_No__icontains=global_search_query)

        vehicle_ids = Vehicle.objects.filter(vehicle_number__icontains=global_search_query).values_list("id", flat=True)
        if vehicle_ids:
            q_objects |= Q(Vehicle_No__in=vehicle_ids)

        consignee_ids = Consignee.objects.filter(consignee_name__icontains=global_search_query).values_list("id", flat=True)
        if consignee_ids:
            q_objects |= Q(consignee__in=consignee_ids)

        consigner_ids = Consigner.objects.filter(consigner_name__icontains=global_search_query).values_list("id", flat=True)
        if consigner_ids:
            q_objects |= Q(consigner__in=consigner_ids)

        location_ids = Location.objects.filter(location_name__icontains=global_search_query).values_list("id", flat=True)
        if location_ids:
            q_objects |= Q(from_location__in=location_ids) | Q(To_Location__in=location_ids)

        consignment_list = consignment_list.filter(q_objects)

    vehicles_lookup = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees_lookup = {str(c.id): c.consignee_name for c in Consignee.objects.all()}

    formatted_consignments = []

    for c in consignment_list:
        goods = c.goods_items.all()

        total_20kg = sum(float(gi.quantity or 0) for gi in goods if gi.unit == "20kg")
        total_10kg = sum(float(gi.quantity or 0) for gi in goods if gi.unit == "10kg")
        total_builty = sum(float(gi.gi_amount or 0) for gi in goods)

        total = (c.Truck_Freight or 0) + (c.Innam or 0) + (c.Extra_TF or 0)

        formatted_consignments.append({
            "Cn_No": c.Cn_No,
            "Loading_Date": c.Loading_Date.strftime("%Y-%m-%d") if c.Loading_Date else "",
            "Vehicle_No": vehicles_lookup.get(str(c.Vehicle_No.id), str(c.Vehicle_No)),
            "consignee": consignees_lookup.get(str(c.consignee.id), str(c.consignee)),
            "get_20kg_total": total_20kg,
            "get_10kg_total": total_10kg,
            "get_builty_fare": total_builty,
            "Truck_Freight": c.Truck_Freight,
            "Innam": c.Innam,
            "Extra_TF": c.Extra_TF,
            "get_total": total,
            "Advance_Amount": c.Advance_Amount,
            "Balance_Amount": c.Balance_Amount,
        })

    paginator = Paginator(formatted_consignments, 10)
    page_number = int(request.GET.get("page", 1))
    page_obj = paginator.get_page(page_number)

    total_pages = paginator.num_pages
    window_size = 4
    current_window = ceil(page_number / window_size)
    start_page = (current_window - 1) * window_size + 1
    end_page = min(start_page + window_size - 1, total_pages)
    custom_page_range = range(start_page, end_page + 1)

    current_filters = {
        "global_search": global_search_query,
    }

    return render(request, "reports.html", {
        "page_obj": page_obj,
        "query_string": request.GET.urlencode(),
        "current_filters": current_filters,
        "page_range": custom_page_range,
        "current_page": page_number,
        "start_page": start_page,
        "end_page": end_page,
        "total_pages": total_pages,
    })

# ──────────────────────────────────────────────────────────────
# Export to Excel 
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def export_builty_excel(request):
    global_search_query = request.GET.get("global_search", "").strip()

    # Fetch consignment data with related goods_items
    consignment_list = Consignment.objects.prefetch_related('goods_items').order_by("CNID")

    # Apply global search filter if present
    if global_search_query:
        consignment_list = consignment_list.filter(
            Q(Cn_No__icontains=global_search_query) |
            Q(Vehicle_No__Vehicle_Number__icontains=global_search_query) |
            Q(consignee__name__icontains=global_search_query)
        )

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger Report"

    # Define headers
    headers = [
        "Bill No", "Loading Date", "Vehicle No", "Consignee",
        "20kg Qty", "10kg Qty", "Builty Fare", "Truck Fare",
        "Innam", "Extra TF", "Total", "Advance Amount", "To Pay"
    ]
    ws.append(headers)

    # Add data rows
    for c in consignment_list:
        goods = c.goods_items.all()
        total_20kg = sum(float(gi.quantity or 0) for gi in goods if gi.unit == "20kg")
        total_10kg = sum(float(gi.quantity or 0) for gi in goods if gi.unit == "10kg")
        total_builty = sum(float(gi.gi_amount or 0) for gi in goods)
        total = (c.Truck_Freight or 0) + (c.Innam or 0) + (c.Extra_TF or 0)

        ws.append([
            c.Cn_No,
            c.Loading_Date.strftime("%Y-%m-%d") if c.Loading_Date else "",
            str(c.Vehicle_No),
            str(c.consignee),
            total_20kg,
            total_10kg,
            total_builty,
            c.Truck_Freight or 0,
            c.Innam or 0,
            c.Extra_TF or 0,
            total,
            c.Advance_Amount or 0,
            c.Balance_Amount or 0,
        ])

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ledger_report.xlsx'
    wb.save(response)
    return response

# ───────────────────────────────────────────────────────────────
# PDF Report Export 
# ─────────────────────────────────────────────────────────────
def export_builty_pdf(request):
    global_search_query = request.GET.get("global_search", "").strip()
    consignment_list = Consignment.objects.prefetch_related("goods_items").order_by("CNID")

    if global_search_query:
        q_objects = Q(Cn_No__icontains=global_search_query)

        vehicle_ids = Vehicle.objects.filter(vehicle_number__icontains=global_search_query).values_list("id", flat=True)
        if vehicle_ids:
            q_objects |= Q(Vehicle_No__in=vehicle_ids)

        consignee_ids = Consignee.objects.filter(consignee_name__icontains=global_search_query).values_list("id", flat=True)
        if consignee_ids:
            q_objects |= Q(consignee__in=consignee_ids)

        consigner_ids = Consigner.objects.filter(consigner_name__icontains=global_search_query).values_list("id", flat=True)
        if consigner_ids:
            q_objects |= Q(consigner__in=consigner_ids)

        location_ids = Location.objects.filter(location_name__icontains=global_search_query).values_list("id", flat=True)
        if location_ids:
            q_objects |= Q(from_location__in=location_ids) | Q(To_Location__in=location_ids)

        consignment_list = consignment_list.filter(q_objects)

    vehicles_lookup = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees_lookup = {str(c.id): c.consignee_name for c in Consignee.objects.all()}

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="Daily_report.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Company Header Block
    company_style = styles['Title'].clone('CompanyStyle')
    company_style.fontSize = 16
    company_style.textColor = colors.HexColor('#1a365d')
    company_style.alignment = TA_CENTER

    subtitle_style = styles['Normal'].clone('SubtitleStyle')
    subtitle_style.fontSize = 10
    subtitle_style.alignment = TA_CENTER

    address_style = styles['Normal'].clone('AddressStyle')
    address_style.fontSize = 9
    address_style.alignment = TA_CENTER

    logo_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
    ]

    logo_cell = None
    for logo_path in logo_paths:
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3 * cm, height=2.5 * cm)
            logo.hAlign = 'RIGHT'
            logo_cell = logo
            break
    else:
        logo_cell = Paragraph("DHANLAXMI", company_style)

    address_block = [
        Paragraph("<b>DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS</b>", company_style),
        Paragraph("<b>IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING</b>", subtitle_style),
        Paragraph("18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304", address_style),
        Paragraph("Mob: +91 9830357491 | Email: kakashri7314@gmail.com", address_style),
    ]

    header_table = Table([[address_block, logo_cell]], colWidths=[16 * cm, 5 * cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))

    # Title
    title_style = styles['Title'].clone('TitleStyle')
    title_style.fontSize = 14
    title_style.textColor = colors.HexColor('#6a1b9a')
    title_style.alignment = TA_CENTER
    elements.append(Paragraph("<b>Daily Report</b>", title_style))
    elements.append(Spacer(1, 12))

    # Table Data
    data = [[
        "Bill No", "Loading Date", "Vehicle No", "Consignee",
        "20kg Qty", "10kg Qty", "Builty Fare", "Truck Fare",
        "Innam", "Extra TF", "Total", "Advance Amount", "To Pay"
    ]]

    for c in consignment_list:
        goods = c.goods_items.all()

        qty_20kg = sum(float(gi.quantity or 0) for gi in goods if (gi.unit or "").strip().upper() == "20KG")
        qty_10kg = sum(float(gi.quantity or 0) for gi in goods if (gi.unit or "").strip().upper() == "10KG")
        builty_fare = sum(float(gi.gi_amount or 0) for gi in goods)

        truck_freight = float(c.Truck_Freight or 0)
        innam = float(c.Innam or 0)
        extra_tf = float(c.Extra_TF or 0)
        advance = float(c.Advance_Amount or 0)
        balance = float(c.Balance_Amount or 0)

        total = truck_freight + innam + extra_tf

        data.append([
            c.Cn_No,
            c.Loading_Date.strftime("%d-%m-%Y") if c.Loading_Date else "",
            vehicles_lookup.get(str(c.Vehicle_No_id), str(c.Vehicle_No_id)),
            consignees_lookup.get(str(c.consignee_id), str(c.consignee_id)),
            f"{qty_20kg:.2f}".rstrip('0').rstrip('.'),
            f"{qty_10kg:.2f}".rstrip('0').rstrip('.'),
            f"{builty_fare:.2f}".rstrip('0').rstrip('.'),
            f"{truck_freight:.2f}".rstrip('0').rstrip('.'),
            f"{innam:.2f}".rstrip('0').rstrip('.'),
            f"{extra_tf:.2f}".rstrip('0').rstrip('.'),
            f"{total:.2f}".rstrip('0').rstrip('.'),
            f"{advance:.2f}".rstrip('0').rstrip('.'),
            f"{balance:.2f}".rstrip('0').rstrip('.'),
        ])


    col_widths = [
        2.2*cm, 2.5*cm, 2.5*cm, 3.0*cm,
        2.0*cm, 2.0*cm, 2.2*cm, 2.2*cm,
        2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm,
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Header row bold
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),      # Body rows normal
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    # Footer Timestamp
    footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    footer_style = styles['Normal'].clone('FooterStyle')
    footer_style.fontSize = 8
    footer_style.textColor = colors.grey
    footer_style.alignment = TA_CENTER
    elements.append(Paragraph(footer_text, footer_style))

    doc.build(elements)
    return response

# ──────────────────────────────────────────────────────────────
# RECORD PAYMENT VIEW
# ──────────────────────────────────────────────────────────────

@login_required(login_url="login")
def record_payment_view(request):
    consignees = Consignee.objects.all()
    consigners = Consigner.objects.all()

    parties = []
    for c in consignees:
        parties.append({"name": c.consignee_name, "type": "Consignee"})
    for c in consigners:
        parties.append({"name": c.consigner_name, "type": "Consigner"})

    context = {
        "parties": parties,
        "parties_json": json.dumps(parties),
    }

    return render(request, "payment_record.html", context)


def get_payment_data(request):
    party_name = request.GET.get("consignee", "").strip()

    if not party_name:
        return JsonResponse({'success': True, 'payments': [], 'message': 'No party specified.'}, safe=False)

    # Filter consignments where either consignee or consigner matches the name
    consignments = Consignment.objects.filter(
        Balance_Amount__gt=0
    ).filter(
        models.Q(consignee__consignee_name__iexact=party_name) |
        models.Q(consigner__consigner_name__iexact=party_name)
    ).order_by("CNID")

    data = []
    for cons in consignments:
        party_type = "Consignee" if cons.consignee.consignee_name.lower() == party_name.lower() else "Consigner"
        data.append({
            "cnid": cons.CNID,
            "consignee": cons.consignee.consignee_name,
            "balance_amount": str(cons.Balance_Amount),
            "party_type": party_type
        })

    return JsonResponse({
        'success': True,
        'payments': data,
        'message': f"{len(data)} records found."
    }, safe=False)


def record_payment(request):
    try:
        data = json.loads(request.body)
        party = data.get('consignee')
        payments = data.get('payments', [])
        payment_date_str = data.get('payment_date')

        if not party or not payments:
            return JsonResponse({'success': False, 'message': 'Missing consignee or payments.'}, status=400)

        # Parse payment_date correctly using datetime.datetime.strptime
        payment_date = None
        if payment_date_str:
            try:
                payment_date = datetime.datetime.strptime(payment_date_str, "%Y-%m-%d").date()
            except Exception as e:
                return JsonResponse({'success': False, 'message': f"Invalid payment date format: {e}"}, status=400)

        with transaction.atomic():
            for item in payments:
                cnid = item.get('cnid')
                paid = item.get('paid_amount')
                balance = item.get('balance_amount')

                if not cnid or paid is None or balance is None:
                    return JsonResponse({'success': False, 'message': 'Incomplete payment data.'}, status=400)

                try:
                    paid = float(paid)
                    balance = float(balance)
                except (ValueError, TypeError):
                    return JsonResponse({'success': False, 'message': f'Invalid amount format for CNID {cnid}.'}, status=400)

                try:
                    cons = Consignment.objects.get(CNID=cnid)
                    cons.Balance_Amount = balance
                    cons.save()

                    PaymentRecord.objects.create(
                        cnid=cons.CNID,
                        consignee=party,
                        total_fare=cons.total_fare,
                        paid_amount=paid,
                        balance_amount=balance,
                        payment_date=payment_date
                    )
                except Consignment.DoesNotExist:
                    continue  # Skip if CNID not found

        return JsonResponse({'success': True, 'message': 'Payments recorded successfully.'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON format.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ──────────────────────────────────────────────────────────────
# Ledger VIEW
# ──────────────────────────────────────────────────────────────


@login_required(login_url="login")
def ledger_view(request):
    global_search_query = request.GET.get("global_search", "").strip()

    # Fetch base queryset ordered by Loading Date
    queryset = Consignment.objects.select_related("Vehicle_No", "consignee").order_by("-Loading_Date")

    # Apply global search (by truck number)
    if global_search_query:
        matching_vehicle_ids = Vehicle.objects.filter(
            vehicle_number__icontains=global_search_query
        ).values_list("id", flat=True)
        queryset = queryset.filter(Vehicle_No__in=matching_vehicle_ids) if matching_vehicle_ids else queryset.none()

    # Add qty_10kg and qty_20kg to each consignment
    for consignment in queryset:
        goods_items = GoodsInfo.objects.filter(consignment=consignment, unit="kg")
        consignment.qty_10kg = goods_items.filter(quantity=10).aggregate(Sum("quantity"))["quantity__sum"] or 0
        consignment.qty_20kg = goods_items.filter(quantity=20).aggregate(Sum("quantity"))["quantity__sum"] or 0

    # Flat list of consignments — no grouping
    ledger_data = list(queryset)

    # Paginate ledger_data – 1 date group per page
    paginator = Paginator(ledger_data, 10)
    page_number = int(request.GET.get("page", 1))
    page_obj = paginator.get_page(page_number)

    # Sliding window pagination (4 pages at a time)
    window_size = 4
    current_window = ceil(page_number / window_size)
    start_page = (current_window - 1) * window_size + 1
    end_page = min(start_page + window_size - 1, paginator.num_pages)
    custom_page_range = range(start_page, end_page + 1)

    # Preserve filters during pagination
    query_params = request.GET.copy()
    query_params.pop("page", None)

    return render(request, "ledger.html", {
        "page_obj": page_obj,
        "page_range": custom_page_range,
        "current_page": page_number,
        "start_page": start_page,
        "end_page": end_page,
        "total_pages": paginator.num_pages,
        "query_string": query_params.urlencode(),
        "global_search": global_search_query,
    })

# ────────────────────────────────────────────────────────────────
# Ledger Excel Export
# ───────────────────────────────────────────────────────────────
@login_required(login_url='login')
def export_ledger_excel(request):
    global_search = request.GET.get("global_search", "").strip()

    entries = Consignment.objects.select_related('Vehicle_No', 'consignee').order_by('-Loading_Date')
    if global_search:
        matching_vehicle_ids = Vehicle.objects.filter(
            vehicle_number__icontains=global_search
        ).values_list("id", flat=True)
        entries = entries.filter(Vehicle_No__in=matching_vehicle_ids) if matching_vehicle_ids else entries.none()

    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Ledger"

    # --- Header: Company Info ---
    ws.merge_cells('A1:I1')
    ws.merge_cells('A2:I2')
    ws.merge_cells('A3:I3')
    ws.merge_cells('A4:I4')

    ws['A1'] = "DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS"
    ws['A2'] = "IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING"
    ws['A3'] = "18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304"
    ws['A4'] = "Mob: +91 9830357491 | Email: kakashri7314@gmail.com"

    for i in range(1, 5):
        ws[f"A{i}"].font = Font(bold=True if i == 1 else False, size=11)
        ws[f"A{i}"].alignment = Alignment(horizontal='center')

    # --- Add Logo if available ---
    try:
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
        ]
        for path in logo_paths:
            if os.path.exists(path):
                img = XLImage(path)
                img.width = 100
                img.height = 75
                ws.add_image(img, 'I1')
                break
    except Exception:
        pass

    ws.append([])  # spacer

    # --- Title Row ---
    ws.merge_cells('A6:I6')
    ws['A6'] = "Vehicle Report"
    ws['A6'].font = Font(bold=True, size=14, color="0D47A1")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.append([])

    # --- Table Headers ---
    headers = [
        "Bill No", "Date", "Truck No", "Consignee",
        "Truck Fare", "Innam", "Extra TF", "Advance", "To Pay"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # --- Table Data ---
    for entry in entries:
        ws.append([
            entry.CNID or "",
            entry.Loading_Date.strftime('%d-%m-%Y') if entry.Loading_Date else '',
            entry.Vehicle_No.vehicle_number if entry.Vehicle_No else '',
            entry.consignee.consignee_name if entry.consignee else '',
            float(entry.Truck_Freight or 0),
            float(entry.Innam or 0),
            float(entry.Extra_TF or 0),
            float(entry.Advance_Amount or 0),
            float(entry.Balance_Amount or 0),
        ])

    # --- Auto-fit Columns ---
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)

    # --- Return Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=truck_ledger.xlsx'
    wb.save(response)
    return response
# ────────────────────────────────────────────────────────────────
# Ledger PDF Export
# ────────────────────────────────────────────────────────────────
@login_required(login_url="login")
def export_ledger_pdf(request):
    global_search_query = request.GET.get("global_search", "").strip()
    queryset = Consignment.objects.select_related("Vehicle_No", "consignee").order_by("-Loading_Date")

    if global_search_query:
        vehicle_ids = Vehicle.objects.filter(vehicle_number__icontains=global_search_query).values_list("id", flat=True)
        if vehicle_ids:
            queryset = queryset.filter(Vehicle_No__in=vehicle_ids)
        else:
            queryset = queryset.none()

    vehicles_lookup = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees_lookup = {str(c.id): c.consignee_name for c in Consignee.objects.all()}

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="ledger.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1 * cm
    )

    styles = getSampleStyleSheet()
    elements = []

    # Header
    company_style = styles['Title'].clone('CompanyStyle')
    company_style.fontSize = 16
    company_style.textColor = colors.HexColor('#1a365d')
    company_style.alignment = TA_CENTER

    subtitle_style = styles['Normal'].clone('SubtitleStyle')
    subtitle_style.fontSize = 10
    subtitle_style.alignment = TA_CENTER

    address_style = styles['Normal'].clone('AddressStyle')
    address_style.fontSize = 9
    address_style.alignment = TA_CENTER

    logo_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
    ]

    logo_cell = None
    for logo_path in logo_paths:
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3 * cm, height=2.5 * cm)
            logo.hAlign = 'RIGHT'
            logo_cell = logo
            break
    else:
        logo_cell = Paragraph("DHANLAXMI", company_style)

    address_block = [
        Paragraph("<b>DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS</b>", company_style),
        Paragraph("<b>IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING</b>", subtitle_style),
        Paragraph("18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304", address_style),
        Paragraph("Mob: +91 9830357491 | Email: kakashri7314@gmail.com", address_style),
    ]

    header_table = Table([[address_block, logo_cell]], colWidths=[16 * cm, 5 * cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))

    # Title
    title_style = styles['Title'].clone('TitleStyle')
    title_style.fontSize = 14
    title_style.textColor = colors.HexColor('#6a1b9a')
    title_style.alignment = TA_CENTER
    elements.append(Paragraph("<b>Vehicle Report</b>", title_style))
    elements.append(Spacer(1, 12))

    # Table
    data = [[
        "Bill No", "Date", "Truck No", "Consignee",
        "Truck Fare", "Innam", "Extra TF", "Advance", "To Pay"
    ]]

    for c in queryset:
        data.append([
            c.CNID or "",
            c.Loading_Date.strftime("%d-%m-%Y") if c.Loading_Date else "",
            vehicles_lookup.get(str(c.Vehicle_No_id), ""),
            consignees_lookup.get(str(c.consignee_id), ""),
            f"{float(c.Truck_Freight or 0):.2f}".rstrip('0').rstrip('.'),
            f"{float(c.Innam or 0):.2f}".rstrip('0').rstrip('.'),
            f"{float(c.Extra_TF or 0):.2f}".rstrip('0').rstrip('.'),
            f"{float(c.Advance_Amount or 0):.2f}".rstrip('0').rstrip('.'),
            f"{float(c.Balance_Amount or 0):.2f}".rstrip('0').rstrip('.'),
        ])

    col_widths = [2.5*cm, 2.5*cm, 3.0*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    # Footer Timestamp
    footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    footer_style = styles['Normal'].clone('FooterStyle')
    footer_style.fontSize = 8
    footer_style.textColor = colors.grey
    footer_style.alignment = TA_CENTER
    elements.append(Paragraph(footer_text, footer_style))

    doc.build(elements)
    return response



# ──────────────────────────────────────────────────────────────
# def party_ledger_view(request):
#     search_query = request.GET.get("global_search", "").strip()
#     consignment_list = Consignment.objects.all().order_by("-Loading_Date")

#     if search_query:
#         consignment_list = consignment_list.filter(
#             consignee__consignee_name__icontains=search_query
#         )

#     vehicles = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
#     consignees = {str(c.id): c.consignee_name for c in Consignee.objects.all()}
#     consigners = {str(c.id): c.consigner_name for c in Consigner.objects.all()}

#     entries = []

#     for c in consignment_list:
#         goods_items = GoodsInfo.objects.filter(consignment=c)

#         # ✅ GROUP BY to_party_id to create separate entries
#         unique_parties = goods_items.values('to_party_id', 'from_party_id').distinct()

#         for party_group in unique_parties:
#             to_party_id = party_group['to_party_id']
#             from_party_id = party_group['from_party_id']
            
#             # Filter goods for this specific party combination
#             party_goods = goods_items.filter(
#                 to_party_id=to_party_id,
#                 from_party_id=from_party_id
#             )

#             # Get party names
#             from_party = consignees.get(str(from_party_id), "") if from_party_id else ""
#             to_party = consignees.get(str(to_party_id), "") if to_party_id else ""

#             # Filter using 'unit' not 'quantity'
#             items_10kg = party_goods.filter(unit="10KG")
#             items_20kg = party_goods.filter(unit="20KG")

#             qty_10kg = items_10kg.aggregate(total=Sum("quantity"))["total"] or 0
#             qty_20kg = items_20kg.aggregate(total=Sum("quantity"))["total"] or 0

#             rates_10kg = list(items_10kg.values_list("rate", flat=True))
#             rates_20kg = list(items_20kg.values_list("rate", flat=True))

#             # Fetch fare from gi_amount column (DB value, exact fare)
#             fare_10kg_db = items_10kg.aggregate(total=Sum("gi_amount"))["total"] or 0
#             fare_20kg_db = items_20kg.aggregate(total=Sum("gi_amount"))["total"] or 0

#             # If both 10KG and 20KG items are present → calculate
#             if items_10kg.exists() and items_20kg.exists():
#                 fare_10kg = sum(qty * rate for qty, rate in items_10kg.values_list("quantity", "rate"))
#                 fare_20kg = sum(qty * rate for qty, rate in items_20kg.values_list("quantity", "rate"))
#             else:
#                 fare_10kg = fare_10kg_db
#                 fare_20kg = fare_20kg_db

#             total_fare = fare_10kg + fare_20kg

#             # ✅ Calculate payments for this specific party
#             # Note: You might need to modify this based on how payments are linked to specific parties
#             # Currently assuming payments are at consignment level
#             total_paid = PaymentRecord.objects.filter(
#                 consignment=c,
#                 payment_type__in=["Parcha", "Part Payment"]
#             ).aggregate(total=Sum("paid_amount"))["total"] or 0

#             # ✅ Proportional payment calculation based on fare percentage
#             total_consignment_fare = GoodsInfo.objects.filter(consignment=c).aggregate(
#                 total=Sum("gi_amount"))["total"] or 0
            
#             if total_consignment_fare > 0:
#                 payment_proportion = total_fare / total_consignment_fare
#                 allocated_paid = total_paid * payment_proportion
#             else:
#                 allocated_paid = 0

#             due_amount = total_fare - allocated_paid

#             avg_rate_10kg = round(sum(rates_10kg) / len(rates_10kg), 2) if rates_10kg else 0
#             avg_rate_20kg = round(sum(rates_20kg) / len(rates_20kg), 2) if rates_20kg else 0

#             # ✅ Skip entries where no goods exist for this party
#             if qty_10kg == 0 and qty_20kg == 0:
#                 continue

#             entries.append({
#                 "cn_id": c.CNID,
#                 "Date": c.Loading_Date,
#                 "Vehicle_No": vehicles.get(str(c.Vehicle_No_id), c.Vehicle_No_id),
#                 "consignee": consignees.get(str(c.consignee_id), c.consignee_id),
#                 "from_party": from_party,   
#                 "to_party": to_party,       
#                 "qty_10kg": qty_10kg,
#                 "rate_10kg": avg_rate_10kg,
#                 "qty_20kg": qty_20kg,
#                 "rate_20kg": avg_rate_20kg,
#                 "fare_10kg": fare_10kg,
#                 "fare_20kg": fare_20kg,
#                 "total_fare": total_fare,
#                 "Paid_Amount": round(allocated_paid, 2),
#                 "Due_Amount": round(due_amount, 2),
#             })

#     paginator = Paginator(entries, 10)
#     page = int(request.GET.get("page", 1))
#     page_obj = paginator.get_page(page)

#     total_pages = paginator.num_pages
#     group_size = 4
#     current_window = ceil(page / group_size)
#     start_page = (current_window - 1) * group_size + 1
#     end_page = min(start_page + group_size - 1, total_pages)

#     query_params = request.GET.copy()
#     query_params.pop("page", None)

#     return render(
#         request,
#         "party_ledger.html",
#         {
#             "page_obj": page_obj,
#             "page_range": range(start_page, end_page + 1),
#             "current_page": page,
#             "start_page": start_page,
#             "end_page": end_page,
#             "total_pages": total_pages,
#             "query_string": query_params.urlencode(),
#             "current_filters": {"global_search": search_query},
#         },
#     )



def party_ledger_view(request):
    to_party_search = request.GET.get("to_party_search", "").strip()
    vehicle_search = request.GET.get("vehicle_search", "").strip()

    consignment_list = Consignment.objects.all().order_by("-Loading_Date")

    # if to_party_search:
    #     consignment_list = consignment_list.filter(consignee__consignee_name__icontains=to_party_search)
    if to_party_search:
        matching_to_party_ids = Consignee.objects.filter(
            consignee_name__icontains=to_party_search
        ).values_list("id", flat=True)

        matching_consignment_ids = GoodsInfo.objects.filter(
            to_party_id__in=matching_to_party_ids
        ).values_list("consignment_id", flat=True).distinct()

        consignment_list = consignment_list.filter(CNID__in=matching_consignment_ids)

    if vehicle_search:
        consignment_list = consignment_list.filter(Vehicle_No__vehicle_number__icontains=vehicle_search)

    vehicles = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees = {str(c.id): c.consignee_name for c in Consignee.objects.all()}
    consigners = {str(c.id): c.consigner_name for c in Consigner.objects.all()}

    entries = []

    for c in consignment_list:
        goods_items = GoodsInfo.objects.filter(consignment=c)

        for goods in goods_items:
            unit = (goods.unit or "").lower()
            quantity = goods.quantity or 0
            rate = goods.rate or 0
            gi_amount = goods.gi_amount or 0
            paid_amount = goods.paid_amount or 0
            balance_amount = goods.balance_amount or 0

            from_party = consignees.get(str(goods.from_party_id), "") if goods.from_party_id else ""
            to_party = consignees.get(str(goods.to_party_id), "") if goods.to_party_id else ""

            qty_10kg = quantity if unit == "10kg" else 0
            rate_10kg = rate if unit == "10kg" else 0
            fare_10kg = quantity * rate if unit == "10kg" else 0

            qty_20kg = quantity if unit == "20kg" else 0
            rate_20kg = rate if unit == "20kg" else 0
            fare_20kg = quantity * rate if unit == "20kg" else 0

            entries.append({
                "cn_id": c.CNID,
                "Date": c.Loading_Date,
                "Vehicle_No": vehicles.get(str(c.Vehicle_No_id), c.Vehicle_No_id),
                "consignee": consignees.get(str(c.consignee_id), c.consignee_id),
                "from_party": from_party,
                "to_party": to_party,
                "qty_10kg": qty_10kg,
                "rate_10kg": rate_10kg,
                "qty_20kg": qty_20kg,
                "rate_20kg": rate_20kg,
                "fare_10kg": fare_10kg,
                "fare_20kg": fare_20kg,
                "total_fare": round(gi_amount, 2),
                "Paid_Amount": round(paid_amount, 2),
                "Due_Amount": round(max(balance_amount, 0), 2),
            })

    # Pagination
    paginator = Paginator(entries, 10)
    page = int(request.GET.get("page", 1))
    page_obj = paginator.get_page(page)

    total_pages = paginator.num_pages
    group_size = 4
    current_window = ceil(page / group_size)
    start_page = (current_window - 1) * group_size + 1
    end_page = min(start_page + group_size - 1, total_pages)

    query_params = request.GET.copy()
    query_params.pop("page", None)

    return render(
        request,
        "party_ledger.html",
        {
            "page_obj": page_obj,
            "page_range": range(start_page, end_page + 1),
            "current_page": page,
            "start_page": start_page,
            "end_page": end_page,
            "total_pages": total_pages,
            "query_string": query_params.urlencode(),
            "current_filters": {
                "to_party_search": to_party_search,
                "vehicle_search": vehicle_search,
            }
        },
    )


# def party_ledger_view(request):
#     search_query = request.GET.get("global_search", "").strip()
#     to_party_search = request.GET.get("to_party_search", "").strip()  # ✅ New search input for To Party

#     consignment_list = Consignment.objects.all().order_by("-Loading_Date")

#     if search_query:
#         consignment_list = consignment_list.filter(
#             consignee__consignee_name__icontains=search_query
#         )

#     vehicles = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
#     consignees = {str(c.id): c.consignee_name for c in Consignee.objects.all()}
#     consigners = {str(c.id): c.consigner_name for c in Consigner.objects.all()}

#     entries = []

#     # ✅ Prepare To Party filter list (IDs)
#     matched_to_party_ids = []
#     if to_party_search:
#         matched_to_party_ids = list(
#             Consignee.objects.filter(
#                 consignee_name__icontains=to_party_search
#             ).values_list("id", flat=True)
#         )

#     for c in consignment_list:
#         goods_items = GoodsInfo.objects.filter(consignment=c)

#         # ✅ GROUP BY to_party_id to create separate entries
#         unique_parties = goods_items.values('to_party_id', 'from_party_id').distinct()

#         for party_group in unique_parties:
#             to_party_id = party_group['to_party_id']
#             from_party_id = party_group['from_party_id']
            
#             # ✅ Filter by To Party name if search applied
#             if to_party_search and to_party_id not in matched_to_party_ids:
#                 continue

#             # Filter goods for this specific party combination
#             party_goods = goods_items.filter(
#                 to_party_id=to_party_id,
#                 from_party_id=from_party_id
#             )

#             # Get party names
#             from_party = consignees.get(str(from_party_id), "") if from_party_id else ""
#             to_party = consignees.get(str(to_party_id), "") if to_party_id else ""

#             # Filter using 'unit' not 'quantity'
#             items_10kg = party_goods.filter(unit="10KG")
#             items_20kg = party_goods.filter(unit="20KG")

#             qty_10kg = items_10kg.aggregate(total=Sum("quantity"))["total"] or 0
#             qty_20kg = items_20kg.aggregate(total=Sum("quantity"))["total"] or 0

#             rates_10kg = list(items_10kg.values_list("rate", flat=True))
#             rates_20kg = list(items_20kg.values_list("rate", flat=True))

#             # Fetch fare from gi_amount column (DB value, exact fare)
#             fare_10kg_db = items_10kg.aggregate(total=Sum("gi_amount"))["total"] or 0
#             fare_20kg_db = items_20kg.aggregate(total=Sum("gi_amount"))["total"] or 0

#             # If both 10KG and 20KG items are present → calculate
#             if items_10kg.exists() and items_20kg.exists():
#                 fare_10kg = sum(qty * rate for qty, rate in items_10kg.values_list("quantity", "rate"))
#                 fare_20kg = sum(qty * rate for qty, rate in items_20kg.values_list("quantity", "rate"))
#             else:
#                 fare_10kg = fare_10kg_db
#                 fare_20kg = fare_20kg_db

#             total_fare = fare_10kg + fare_20kg

#             # ✅ Calculate payments for this specific party
#             total_paid = PaymentRecord.objects.filter(
#                 consignment=c,
#                 payment_type__in=["Parcha", "Part Payment"]
#             ).aggregate(total=Sum("paid_amount"))["total"] or 0

#             # ✅ Proportional payment calculation based on fare percentage
#             total_consignment_fare = GoodsInfo.objects.filter(consignment=c).aggregate(
#                 total=Sum("gi_amount"))["total"] or 0
            
#             if total_consignment_fare > 0:
#                 payment_proportion = total_fare / total_consignment_fare
#                 allocated_paid = total_paid * payment_proportion
#             else:
#                 allocated_paid = 0

#             due_amount = total_fare - allocated_paid

#             avg_rate_10kg = round(sum(rates_10kg) / len(rates_10kg), 2) if rates_10kg else 0
#             avg_rate_20kg = round(sum(rates_20kg) / len(rates_20kg), 2) if rates_20kg else 0

#             # ✅ Skip entries where no goods exist for this party
#             if qty_10kg == 0 and qty_20kg == 0:
#                 continue

#             entries.append({
#                 "cn_id": c.CNID,
#                 "Date": c.Loading_Date,
#                 "Vehicle_No": vehicles.get(str(c.Vehicle_No_id), c.Vehicle_No_id),
#                 "consignee": consignees.get(str(c.consignee_id), c.consignee_id),
#                 "from_party": from_party,   
#                 "to_party": to_party,       
#                 "qty_10kg": qty_10kg,
#                 "rate_10kg": avg_rate_10kg,
#                 "qty_20kg": qty_20kg,
#                 "rate_20kg": avg_rate_20kg,
#                 "fare_10kg": fare_10kg,
#                 "fare_20kg": fare_20kg,
#                 "total_fare": total_fare,
#                 "Paid_Amount": round(allocated_paid, 2),
#                 "Due_Amount": round(due_amount, 2),
#             })

#     paginator = Paginator(entries, 10)
#     page = int(request.GET.get("page", 1))
#     page_obj = paginator.get_page(page)

#     total_pages = paginator.num_pages
#     group_size = 4
#     current_window = ceil(page / group_size)
#     start_page = (current_window - 1) * group_size + 1
#     end_page = min(start_page + group_size - 1, total_pages)

#     query_params = request.GET.copy()
#     query_params.pop("page", None)

#     return render(
#         request,
#         "party_ledger.html",
#         {
#             "page_obj": page_obj,
#             "page_range": range(start_page, end_page + 1),
#             "current_page": page,
#             "start_page": start_page,
#             "end_page": end_page,
#             "total_pages": total_pages,
#             "query_string": query_params.urlencode(),
#             "current_filters": {
#                 "global_search": search_query,
#                 "to_party_search": to_party_search,  # ✅ Include in context
#             },
#         },
#     )

# ──────────────────────────────────────────────────────────────
import logging
logger = logging.getLogger(__name__)
@login_required
def get_payment_history(request):
    cn_id = request.GET.get("cn_id")
    if not cn_id:
        return JsonResponse({"error": "Missing consignment ID"}, status=400)

    records = PaymentRecord.objects.filter(consignment_id=cn_id).order_by("payment_date")
    logger.debug(f"Found {records.count()} payment records for CN {cn_id}")

    data = []
    total_paid = 0
    latest_balance = 0
    for r in records:
        total_paid += r.paid_amount
        latest_balance = r.balance_amount
        data.append({
            "paid_amount": str(r.paid_amount),
            "payment_date": r.payment_date.strftime("%Y-%m-%d"),
            "payment_type": r.payment_type,
            "payment_mode": r.payment_mode,
            "remark": r.remark or "",
        })

    return JsonResponse({
        "history": data,
        "total_paid": str(total_paid),
        "latest_balance": str(latest_balance),
    })
# ──────────────────────────────────────────────────────────────
# Export to Excel 
# ──────────────────────────────────────────────────────────────

@login_required(login_url="login")
def export_party_ledger_excel(request):
    # Initialize Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Goods_report"

    # Header: Company Info
    ws.merge_cells('A1:O1')
    ws.merge_cells('A2:O2')
    ws.merge_cells('A3:O3')
    ws.merge_cells('A4:O4')

    ws['A1'] = "DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS"
    ws['A2'] = "IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING"
    ws['A3'] = "18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304"
    ws['A4'] = "Mob: +91 9830357491 | Email: kakashri7314@gmail.com"

    for i in range(1, 5):
        ws[f"A{i}"].font = Font(bold=(i == 1), size=11)
        ws[f"A{i}"].alignment = Alignment(horizontal='center')

    # Add Logo if Available
    try:
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
        ]
        for path in logo_paths:
            if os.path.exists(path):
                img = XLImage(path)
                img.width = 100
                img.height = 75
                ws.add_image(img, 'O1')
                break
    except Exception:
        pass

    ws.append([])  # Spacer Row

    # Title Row
    ws.merge_cells('A6:O6')
    ws['A6'] = "Goods Report"
    ws['A6'].font = Font(bold=True, size=14, color="0D47A1")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.append([])

    # Table Headers
    headers = [
        "CN No", "Date", "Truck No", "Consignee", "From Party", "To Party",
        "10KG Qty", "10KG Rate", "10KG Fare",
        "20KG Qty", "20KG Rate", "20KG Fare",
        "Total Fare", "Paid", "Due"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='B3E5FC', end_color='B3E5FC', fill_type='solid')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # --- Filters from Request ---
    consignee_query = request.GET.get("consignee_search", "").strip()
    to_party_query = request.GET.get("to_party_search", "").strip()
    vehicle_query = request.GET.get("vehicle_search", "").strip()

    vehicles = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees = {c.id: c.consignee_name for c in Consignee.objects.all()}

    # Base QuerySet
    goods_list = GoodsInfo.objects.select_related("consignment").order_by("-created_at")

    # Apply Filters
    if consignee_query:
        goods_list = goods_list.filter(consignment__consignee__consignee_name__icontains=consignee_query)
    if to_party_query:
        goods_list = goods_list.filter(to_party__consignee_name__icontains=to_party_query)
    if vehicle_query:
        goods_list = goods_list.filter(consignment__Vehicle_No__vehicle_number__icontains=vehicle_query)

    # --- Populate Table Data ---
    for gi in goods_list:
        c = gi.consignment
        vehicle_no = vehicles.get(str(c.Vehicle_No_id), c.Vehicle_No_id)
        consignee_name = consignees.get(c.consignee_id, c.consignee_id)
        from_party = consignees.get(gi.from_party_id, gi.from_party_id)
        to_party = consignees.get(gi.to_party_id, gi.to_party_id)

        gi_amount = (gi.quantity or 0) * (gi.rate or 0)

        total_paid = PaymentRecord.objects.filter(
            consignment=c, payment_type__in=["Parcha", "Part Payment"]
        ).aggregate(total=Sum("paid_amount"))["total"] or 0

        due_amount = gi_amount - total_paid

        # Handle Unit-based Split
        unit = (gi.unit or "").strip().upper()
        qty_10 = rate_10 = fare_10 = qty_20 = rate_20 = fare_20 = 0

        if unit == "10KG":
            qty_10, rate_10, fare_10 = gi.quantity or 0, gi.rate or 0, gi_amount
        elif unit == "20KG":
            qty_20, rate_20, fare_20 = gi.quantity or 0, gi.rate or 0, gi_amount

        row = [
            c.Cn_No,
            c.Loading_Date.strftime('%d-%m-%Y') if c.Loading_Date else '',
            vehicle_no,
            consignee_name,
            from_party,
            to_party,
            f"{qty_10:.2f}".rstrip('0').rstrip('.') if qty_10 else "0",
            f"{rate_10:.2f}".rstrip('0').rstrip('.') if rate_10 else "0",
            f"{fare_10:.2f}".rstrip('0').rstrip('.') if fare_10 else "0",
            f"{qty_20:.2f}".rstrip('0').rstrip('.') if qty_20 else "0",
            f"{rate_20:.2f}".rstrip('0').rstrip('.') if rate_20 else "0",
            f"{fare_20:.2f}".rstrip('0').rstrip('.') if fare_20 else "0",
            f"{gi_amount:.2f}".rstrip('0').rstrip('.') if gi_amount else "0",
            f"{total_paid:.2f}".rstrip('0').rstrip('.') if total_paid else "0",
            f"{due_amount:.2f}".rstrip('0').rstrip('.') if due_amount else "0"
        ]
        ws.append(row)

    # Auto-fit Columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 25)

    # Return Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Goods_report.xlsx'
    wb.save(response)
    return response
# ──────────────────────────────────────────────────────────────
# Export to PDF 
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def export_party_ledger_pdf(request):
    to_party_search = request.GET.get("to_party_search", "").strip()
    vehicle_search = request.GET.get("vehicle_search", "").strip()
    global_search = request.GET.get("global_search", "").strip()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=Party_Ledger_Report.pdf"
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1 * cm, rightMargin=1 * cm,
        topMargin=1.5 * cm, bottomMargin=1 * cm
    )
    styles = getSampleStyleSheet()
    elements = []

    # ─── Header Styles ────────────────────────────
    company_style = styles['Title'].clone('CompanyStyle')
    company_style.fontSize = 16
    company_style.textColor = colors.HexColor('#1a365d')
    company_style.alignment = TA_CENTER

    subtitle_style = styles['Normal'].clone('SubtitleStyle')
    subtitle_style.fontSize = 10
    subtitle_style.alignment = TA_CENTER

    address_style = styles['Normal'].clone('AddressStyle')
    address_style.fontSize = 9
    address_style.alignment = TA_CENTER

    # ─── Logo Block ───────────────────────────────
    logo_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
    ]

    logo_cell = None
    for logo_path in logo_paths:
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3 * cm, height=2.5 * cm)
            logo.hAlign = 'RIGHT'
            logo_cell = logo
            break
    else:
        logo_cell = Paragraph("DHANLAXMI", company_style)

    address_block = [
        Paragraph("<b>DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS</b>", company_style),
        Paragraph("<b>IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING</b>", subtitle_style),
        Paragraph("18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304", address_style),
        Paragraph("Mob: +91 9830357491 | Email: kakashri7314@gmail.com", address_style),
    ]

    header_table = Table([[address_block, logo_cell]], colWidths=[16 * cm, 5 * cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))

    # ─── Title ────────────────────────────────────
    title_style = styles['Title'].clone('TitleStyle')
    title_style.fontSize = 14
    title_style.textColor = colors.HexColor('#6a1b9a')
    title_style.alignment = TA_CENTER
    elements.append(Paragraph("<b>Goods Report</b>", title_style))
    elements.append(Spacer(1, 12))

    # ─── Table Header ─────────────────────────────
    data = [[
        "CN No", "Date", "Truck No", "Consignee", "From Party", "To Party",
        "10KG Qty", "10KG Rate", "10KG Fare",
        "20KG Qty", "20KG Rate", "20KG Fare",
        "Total Fare", "Paid", "Due"
    ]]

    # ─── Lookup Dictionaries ──────────────────────
    vehicles = {str(v.id): v.vehicle_number for v in Vehicle.objects.all()}
    consignees = {c.id: c.consignee_name for c in Consignee.objects.all()}

    goods_list = GoodsInfo.objects.select_related("consignment").order_by("-created_at")

    # ─── Apply Filters ────────────────────────────
    if global_search:
        consignment_ids = Consignment.objects.filter(
            Q(Cn_No__icontains=global_search) |
            Q(Vehicle_No__vehicle_number__icontains=global_search) |
            Q(consignee__consignee_name__icontains=global_search) |
            Q(consigner__consigner_name__icontains=global_search)
        ).values_list('id', flat=True)

        goods_list = goods_list.filter(
            Q(consignment_id__in=consignment_ids) |
            Q(from_party__consignee_name__icontains=global_search) |
            Q(to_party__consignee_name__icontains=global_search)
        )
    else:
        if to_party_search:
            matching_to_party_ids = Consignee.objects.filter(
                consignee_name__icontains=to_party_search
            ).values_list("id", flat=True)

            matching_consignment_ids = GoodsInfo.objects.filter(
                to_party_id__in=matching_to_party_ids
            ).values_list("consignment_id", flat=True).distinct()

            goods_list = goods_list.filter(consignment_id__in=matching_consignment_ids)

        if vehicle_search:
            goods_list = goods_list.filter(consignment__Vehicle_No__vehicle_number__icontains=vehicle_search)

    # ─── Data Rows ────────────────────────────────
    for gi in goods_list:
        c = gi.consignment
        vehicle_no = vehicles.get(str(c.Vehicle_No_id), c.Vehicle_No_id)
        consignee_name = consignees.get(c.consignee_id, c.consignee_id)
        from_party = consignees.get(gi.from_party_id, gi.from_party_id)
        to_party = consignees.get(gi.to_party_id, gi.to_party_id)

        quantity = gi.quantity or 0
        rate = gi.rate or 0
        gi_amount = quantity * rate

        paid_amount = gi.paid_amount or 0
        due_amount = gi.balance_amount or 0

        unit = (gi.unit or "").strip().upper()
        qty_10 = rate_10 = fare_10 = qty_20 = rate_20 = fare_20 = 0

        if unit == "10KG":
            qty_10, rate_10, fare_10 = quantity, rate, gi_amount
        elif unit == "20KG":
            qty_20, rate_20, fare_20 = quantity, rate, gi_amount

        row = [
            c.Cn_No, c.Loading_Date.strftime("%d-%m-%Y") if c.Loading_Date else "",
            vehicle_no, consignee_name, from_party, to_party,
            f"{qty_10:.2f}".rstrip('0').rstrip('.') if qty_10 else "0",
            f"{rate_10:.2f}".rstrip('0').rstrip('.') if rate_10 else "0",
            f"{fare_10:.2f}".rstrip('0').rstrip('.') if fare_10 else "0",
            f"{qty_20:.2f}".rstrip('0').rstrip('.') if qty_20 else "0",
            f"{rate_20:.2f}".rstrip('0').rstrip('.') if rate_20 else "0",
            f"{fare_20:.2f}".rstrip('0').rstrip('.') if fare_20 else "0",
            f"{gi_amount:.2f}".rstrip('0').rstrip('.') if gi_amount else "0",
            f"{paid_amount:.2f}".rstrip('0').rstrip('.') if paid_amount else "0",
            f"{due_amount:.2f}".rstrip('0').rstrip('.') if due_amount else "0"
        ]
        data.append([str(cell)[:15] for cell in row])

    # ─── Table Rendering ──────────────────────────
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    # ─── Footer Timestamp ─────────────────────────
    footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    footer_style = styles['Normal'].clone('FooterStyle')
    footer_style.fontSize = 8
    footer_style.textColor = colors.grey
    footer_style.alignment = TA_CENTER
    elements.append(Paragraph(footer_text, footer_style))

    # ─── Build PDF ────────────────────────────────
    doc.build(elements)
    return response

# ──────────────────────────────────────────────────────────────
# Print Payment History View
# ──────────────────────────────────────────────────────────────


# ______________________________________________________________
# Print Bill View
# ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def print_bill(request, cnid):
    # Step 1: Fetch the consignment(s) for the given CNID
    consignments = Consignment.objects.filter(CNID=cnid)
    if not consignments.exists():
        return HttpResponse("No data found for CNID", status=404)

    all_goods = []

    for consignment in consignments:
        goods_items = GoodsInfo.objects.filter(consignment=consignment)
        for item in goods_items:
            all_goods.append({
                'cnid': consignment.CNID,
                'truck_no': consignment.Vehicle_No.vehicle_number if consignment.Vehicle_No else '',
                'from_party': item.from_party,
                'to_party': item.to_party,
                'qty': item.quantity,
                'unit': item.unit,
                'rate': item.rate,
                'gi_amount': item.gi_amount,
                'loading_date': consignment.Loading_Date.strftime("%d-%m-%Y") if consignment.Loading_Date else '',
                'consigner_name': consignment.consigner.consigner_name if consignment.consigner else '',
                'consignee_name': consignment.consignee.consignee_name if consignment.consignee else '',
            })

    # Step 2: Group the items by 'to_party' to generate separate bills
    grouped = defaultdict(list)
    for item in all_goods:
        grouped[item['to_party']].append(item)

    # Step 3: Prepare bill data structure to pass to the template
    bill_data = []
    unique_to_parties = set()
    for to_party, items in grouped.items():
        first = items[0]
        total = sum(i['gi_amount'] for i in items)

        # Convert total amount to words
        total_in_words = num2words(total, lang='en_IN').title() + " Only"

        bill_data.append({
            'cnid': first['cnid'],
            'truck_no': first['truck_no'],
            'consigner': first['consigner_name'].split()[0] if first['consigner_name'] else '',
            'consignee': first['consignee_name'],
            'items': items,
            'total_fare': total,
            'total_fare_words': total_in_words,
            'loading_date': first['loading_date'],
        })
    
        unique_to_parties.add(to_party)

    # Step 4: Render template with grouped data and unique To Parties for filter
    return render(request, 'print_bill.html', {
        'bill_data': bill_data,
        'unique_to_parties': sorted(str(party) for party in unique_to_parties),  # pass to template
    })  
    
#_____________________________________________________________________________


# @csrf_exempt
# @login_required(login_url="login")
# def cash_book_view(request):
#     if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
#         data = json.loads(request.body)

#         party = data.get("party")
#         paid_amount = Decimal(str(data.get("paid_amount", "0")))
#         payment_date = data.get("payment_date")
#         payment_type = data.get("payment_type")
#         payment_mode = data.get("payment_mode")
#         vehicle_no = data.get("vehicle_no")
#         loading_date = data.get("date")

#         consignee_id = Consignee.objects.filter(consignee_name=party).values_list("id", flat=True).first()
#         consigner_id = Consigner.objects.filter(consigner_name=party).values_list("id", flat=True).first()
#         valid_party_ids = list(filter(None, [consignee_id, consigner_id]))

#         goods_list = GoodsInfo.objects.filter(
#             to_party__in=valid_party_ids,
#             consignment__Vehicle_No__vehicle_number=vehicle_no,
#             consignment__Loading_Date=loading_date,
#         ).order_by("GI_ID")

#         if not goods_list.exists():
#             return JsonResponse({"error": "GoodsInfo not found"}, status=404)

#         updated_goods_ids = []

#         for goods in goods_list:
#             if paid_amount <= 0:
#                 break

#             total_paid_before = PaymentRecord.objects.filter(goods_info_ids__contains=[goods.id]).aggregate(
#                 total=Sum('paid_amount'))['total'] or Decimal("0.00")

#             remaining_balance = goods.gi_amount - total_paid_before

#             if remaining_balance <= 0:
#                 continue

#             pay_now = min(paid_amount, remaining_balance)
#             balance_after_payment = remaining_balance - pay_now

#             updated_goods_ids.append(goods.id)

#             # Create PaymentRecord (one per goods)
#             PaymentRecord.objects.create(
#                 goods_info_ids=[goods.id],
#                 party=party,
#                 balance_amount=balance_after_payment,
#                 paid_amount=pay_now,
#                 payment_type=payment_type,
#                 payment_mode=payment_mode,
#                 payment_date=payment_date,
#                 remark="remark",
#                 created_at=timezone.now(),
#             )

#             # Update GoodsInfo
#             goods.paid_amount = total_paid_before + pay_now
#             goods.balance_amount = goods.gi_amount - goods.paid_amount
#             goods.party_payment = payment_date
#             goods.save(update_fields=["paid_amount", "balance_amount", "party_payment"])

#             paid_amount -= pay_now

#         return JsonResponse({"success": True})

#     # Get current balance for selected party, vehicle, and date
#     elif request.GET.get("to_party") and request.GET.get("vehicle_no") and request.GET.get("date"):
#         party_name = request.GET.get("to_party")
#         vehicle_no = request.GET.get("vehicle_no")
#         date = request.GET.get("date")

#         consignee_id = Consignee.objects.filter(consignee_name=party_name).values_list("id", flat=True).first()
#         consigner_id = Consigner.objects.filter(consigner_name=party_name).values_list("id", flat=True).first()
#         valid_party_ids = list(filter(None, [consignee_id, consigner_id]))

#         goods_list = GoodsInfo.objects.filter(
#             to_party__in=valid_party_ids,
#             consignment__Vehicle_No__vehicle_number=vehicle_no,
#             consignment__Loading_Date=date,
#         )

#         if goods_list.exists():
#             total_gi_amount = goods_list.aggregate(total=Sum('gi_amount'))['total'] or 0
#             total_paid = 0
#             for g in goods_list:
#                 paid = PaymentRecord.objects.filter(goods_info_ids__contains=[g.id]).aggregate(
#                     total=Sum('paid_amount'))['total'] or 0
#                 total_paid += paid
#             balance = total_gi_amount - total_paid

#             return JsonResponse({
#                 "gi_amount": float(total_gi_amount),
#                 "paid_amount": float(total_paid),
#                 "balance_amount": round(float(balance), 2)
#             })
#         else:
#             return JsonResponse({"gi_amount": 0, "paid_amount": 0, "balance_amount": 0})

#     # Get vehicle numbers for selected date
#     elif request.GET.get("date") and not request.GET.get("vehicle_no"):
#         date = request.GET.get("date")
#         vehicles = Consignment.objects.filter(Loading_Date=date).select_related("Vehicle_No").values_list(
#             "Vehicle_No__vehicle_number", flat=True).distinct()
#         return JsonResponse(list(vehicles), safe=False)

#     # Get TO PARTY names for selected vehicle
#     elif request.GET.get("vehicle_no"):
#         vehicle_no = request.GET.get("vehicle_no")

#         to_party_ids = GoodsInfo.objects.filter(consignment__Vehicle_No__vehicle_number=vehicle_no).values_list(
#             "to_party", flat=True).distinct()

#         names = set()
#         for pid in to_party_ids:
#             if not pid:
#                 continue
#             if str(pid).isdigit():
#                 pid = int(pid)
#                 name = (
#                     Consignee.objects.filter(id=pid).values_list("consignee_name", flat=True).first()
#                     or Consigner.objects.filter(id=pid).values_list("consigner_name", flat=True).first()
#                 )
#                 if name:
#                     names.add(name)

#         return JsonResponse([{"name": name} for name in names], safe=False)

#     return render(request, "cash_book.html")

# @csrf_exempt
# @login_required(login_url="login")
# def cash_book_view(request):
#     if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
#         data = json.loads(request.body)

#         party = data.get("party")
#         paid_amount = Decimal(str(data.get("paid_amount", "0")))
#         payment_date = data.get("payment_date")
#         payment_type = data.get("payment_type")
#         payment_mode = data.get("payment_mode")
#         vehicle_no = data.get("vehicle_no")
#         loading_date = data.get("date")

#         # Find party ID from either Consignee or Consigner
#         consignee_id = Consignee.objects.filter(consignee_name=party).values_list("id", flat=True).first()
#         consigner_id = Consigner.objects.filter(consigner_name=party).values_list("id", flat=True).first()
#         valid_party_ids = list(filter(None, [consignee_id, consigner_id]))

#         # Find all matching GoodsInfo records ordered by GI_ID
#         goods_list = GoodsInfo.objects.filter(
#             to_party__in=valid_party_ids,
#             consignment__Vehicle_No__vehicle_number=vehicle_no,
#             consignment__Loading_Date=loading_date,
#         ).order_by("GI_ID")

#         if not goods_list.exists():
#             return JsonResponse({"error": "GoodsInfo not found"}, status=404)

#         for goods in goods_list:
#             if paid_amount <= 0:
#                 break

#             total_paid_before = PaymentRecord.objects.filter(goods_info=goods).aggregate(
#                 total=Sum('paid_amount'))['total'] or Decimal("0.00")

#             remaining_balance = goods.gi_amount - total_paid_before

#             if remaining_balance <= 0:
#                 continue

#             pay_now = min(paid_amount, remaining_balance)
#             balance_after_payment = remaining_balance - pay_now

#             PaymentRecord.objects.create(
#                 goods_info=goods,
#                 consignment=goods.consignment,
#                 party=party,
#                 balance_amount=balance_after_payment,
#                 paid_amount=pay_now,
#                 payment_type=payment_type,
#                 payment_mode=payment_mode,
#                 payment_date=payment_date,
#                 remark="remark",
#                 created_at=timezone.now(),
#             )

#             goods.paid_amount = total_paid_before + pay_now
#             goods.balance_amount = goods.gi_amount - goods.paid_amount
#             goods.party_payment = payment_date
#             goods.save(update_fields=["paid_amount", "balance_amount", "party_payment"])

#             paid_amount -= pay_now

#         return JsonResponse({"success": True})

#     # Get current balance for selected party, vehicle, and date
#     elif request.GET.get("to_party") and request.GET.get("vehicle_no") and request.GET.get("date"):
#         party_name = request.GET.get("to_party")
#         vehicle_no = request.GET.get("vehicle_no")
#         date = request.GET.get("date")

#         consignee_id = Consignee.objects.filter(consignee_name=party_name).values_list("id", flat=True).first()
#         consigner_id = Consigner.objects.filter(consigner_name=party_name).values_list("id", flat=True).first()
#         valid_party_ids = list(filter(None, [consignee_id, consigner_id]))

#         goods_list = GoodsInfo.objects.filter(
#             to_party__in=valid_party_ids,
#             consignment__Vehicle_No__vehicle_number=vehicle_no,
#             consignment__Loading_Date=date,
#         )

#         if goods_list.exists():
#             total_gi_amount = goods_list.aggregate(total=Sum('gi_amount'))['total'] or 0
#             total_paid = PaymentRecord.objects.filter(goods_info__in=goods_list).aggregate(total=Sum('paid_amount'))['total'] or 0
#             balance = total_gi_amount - total_paid

#             return JsonResponse({
#                 "gi_amount": float(total_gi_amount),
#                 "paid_amount": float(total_paid),
#                 "balance_amount": round(float(balance), 2)
#             })
#         else:
#             return JsonResponse({"gi_amount": 0, "paid_amount": 0, "balance_amount": 0})

#     # Get vehicle numbers for selected date
#     elif request.GET.get("date") and not request.GET.get("vehicle_no"):
#         date = request.GET.get("date")
#         vehicles = Consignment.objects.filter(Loading_Date=date).select_related("Vehicle_No").values_list(
#             "Vehicle_No__vehicle_number", flat=True).distinct()
#         return JsonResponse(list(vehicles), safe=False)

#     # Get TO PARTY names for selected vehicle
#     elif request.GET.get("vehicle_no"):
#         vehicle_no = request.GET.get("vehicle_no")

#         to_party_ids = GoodsInfo.objects.filter(consignment__Vehicle_No__vehicle_number=vehicle_no).values_list(
#             "to_party", flat=True).distinct()

#         names = set()
#         for pid in to_party_ids:
#             if not pid:
#                 continue
#             if str(pid).isdigit():
#                 pid = int(pid)
#                 name = (
#                     Consignee.objects.filter(id=pid).values_list("consignee_name", flat=True).first()
#                     or Consigner.objects.filter(id=pid).values_list("consigner_name", flat=True).first()
#                 )
#                 if name:
#                     names.add(name)

#         return JsonResponse([{"name": name} for name in names], safe=False)

#     return render(request, "cash_book.html")
@csrf_exempt
@login_required(login_url="login")
def parcha_view(request):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            data = json.loads(request.body)

            party = data.get("party")
            paid_amount = Decimal(str(data.get("paid_amount", "0")))
            payment_date = data.get("payment_date")
            payment_type = data.get("payment_type")
            payment_mode = data.get("payment_mode")
            vehicle_no = data.get("vehicle_no")
            loading_date = data.get("date")
            remark = data.get("remark", "remark")

            consignee_id = Consignee.objects.filter(consignee_name=party).values_list("id", flat=True).first()
            consigner_id = Consigner.objects.filter(consigner_name=party).values_list("id", flat=True).first()
            valid_party_ids = list(filter(None, [consignee_id, consigner_id]))

            # Get all matching GoodsInfo
            goods_qs = GoodsInfo.objects.filter(
                to_party__in=valid_party_ids,
                consignment__Vehicle_No__vehicle_number=vehicle_no,
                consignment__Loading_Date=loading_date,
            ).select_related("consignment")

            if not goods_qs.exists():
                return JsonResponse({"error": "GoodsInfo not found"}, status=404)

            # Calculate total gi_amount and total previously paid
            total_gi_amount = goods_qs.aggregate(total=Sum("gi_amount"))["total"] or Decimal("0.00")

            total_paid_before = PaymentRecord.objects.filter(goods_info__in=goods_qs).aggregate(
                total=Sum("paid_amount")
            )["total"] or Decimal("0.00")

            balance_after_payment = total_gi_amount - total_paid_before - paid_amount

            # Save payment against the first matching GoodsInfo (you could improve this)
            goods = goods_qs.first()

            PaymentRecord.objects.create(
                goods_info=goods,
                consignment=goods.consignment,
                party=party,
                balance_amount=balance_after_payment,
                paid_amount=paid_amount,
                payment_type=payment_type,
                payment_mode=payment_mode,
                payment_date=payment_date,
                remark=remark,
                created_at=timezone.now(),
            )

            goods.paid_amount = (goods.paid_amount or Decimal("0.00")) + paid_amount
            goods.balance_amount = (goods.gi_amount or Decimal("0.00")) - goods.paid_amount
            goods.party_payment = payment_date
            goods.save(update_fields=["paid_amount", "balance_amount", "party_payment"])

            return JsonResponse({"success": True})
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"error": f"Server error: {str(e)}"}, status=500)

    elif request.GET.get("to_party") and request.GET.get("vehicle_no") and request.GET.get("date"):
        try:
            party_name = request.GET.get("to_party")
            vehicle_no = request.GET.get("vehicle_no")
            date = request.GET.get("date")

            consignee_id = Consignee.objects.filter(consignee_name=party_name).values_list("id", flat=True).first()
            consigner_id = Consigner.objects.filter(consigner_name=party_name).values_list("id", flat=True).first()
            valid_party_ids = list(filter(None, [consignee_id, consigner_id]))

            goods_qs = GoodsInfo.objects.filter(
                to_party__in=valid_party_ids,
                consignment__Vehicle_No__vehicle_number=vehicle_no,
                consignment__Loading_Date=date,
            )

            if goods_qs.exists():
                total_gi = goods_qs.aggregate(total=Sum("gi_amount"))["total"] or Decimal("0.00")
                total_paid = goods_qs.aggregate(total=Sum("paid_amount"))["total"] or Decimal("0.00")
                balance = total_gi - total_paid

                return JsonResponse({
                    "gi_amount": float(total_gi),
                    "paid_amount": float(total_paid),
                    "balance_amount": round(float(balance), 2),
                })
            else:
                return JsonResponse({"gi_amount": 0, "paid_amount": 0, "balance_amount": 0})
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"error": f"Server error: {str(e)}"}, status=500)

    elif request.GET.get("date") and not request.GET.get("vehicle_no"):
        try:
            date = request.GET.get("date")
            vehicles = Consignment.objects.filter(Loading_Date=date).select_related("Vehicle_No").values_list(
                "Vehicle_No__vehicle_number", flat=True
            ).distinct()
            return JsonResponse(list(vehicles), safe=False)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"error": f"Server error: {str(e)}"}, status=500)

    elif request.GET.get("vehicle_no"):
        try:
            vehicle_no = request.GET.get("vehicle_no")

            to_party_ids = GoodsInfo.objects.filter(
                consignment__Vehicle_No__vehicle_number=vehicle_no
            ).values_list("to_party", flat=True).distinct()

            names = set()
            for pid in to_party_ids:
                if not pid:
                    continue
                name = (
                    Consignee.objects.filter(id=pid).values_list("consignee_name", flat=True).first()
                    or Consigner.objects.filter(id=pid).values_list("consigner_name", flat=True).first()
                )
                if name:
                    names.add(name)

            return JsonResponse([{"name": name} for name in names], safe=False)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"error": f"Server error: {str(e)}"}, status=500)

    return render(request, "record_payment.html")


@csrf_exempt
@login_required(login_url="login")
def part_payment_view(request):
    if request.method == "GET" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            party_name = request.GET.get("part_payment_balance")

            if party_name:
                consignee_id = Consignee.objects.filter(
                    consignee_name=party_name
                ).values_list("id", flat=True).first()

                if not consignee_id:
                    return JsonResponse({"balance_amount": 0})

                goods_qs = GoodsInfo.objects.filter(to_party=consignee_id)

                total_balance = Decimal("0.00")

                for g in goods_qs:
                    gi_amt = g.gi_amount or Decimal("0.00")
                    paid_amt = PaymentRecord.objects.filter(goods_info=g).aggregate(
                        total=Sum("paid_amount")
                    )["total"] or Decimal("0.00")

                    total_balance += gi_amt - paid_amt

                return JsonResponse({"balance_amount": float(total_balance)})

            consignee_names = list(Consignee.objects.values_list("consignee_name", flat=True).distinct())
            return JsonResponse(sorted(consignee_names), safe=False)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"error": f"Server error: {str(e)}"}, status=500)

    elif request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            data = json.loads(request.body)
            party = data.get("party")
            paid_amount = Decimal(str(data.get("paid_amount", "0")))
            payment_date = data.get("payment_date")
            payment_mode = data.get("payment_mode")
            remark = data.get("remark", "")

            if not party or paid_amount <= 0 or not payment_date or not payment_mode:
                return JsonResponse({"error": "Missing required fields"}, status=400)

            consignee_id = Consignee.objects.filter(
                consignee_name=party
            ).values_list("id", flat=True).first()

            if not consignee_id:
                return JsonResponse({"error": "Invalid party name."}, status=400)

            goods_list = GoodsInfo.objects.filter(to_party=consignee_id).order_by("GI_ID")
            original_paid_amount = paid_amount
            updated_goods = []

            for goods in goods_list:
                if paid_amount <= 0:
                    break

                gi_amt = goods.gi_amount or Decimal("0.00")

                total_paid_before = PaymentRecord.objects.filter(
                    goods_info=goods
                ).aggregate(
                    total=Sum("paid_amount")
                )['total'] or Decimal("0.00")

                remaining_balance = gi_amt - total_paid_before

                if remaining_balance <= 0:
                    continue

                pay_now = min(paid_amount, remaining_balance)

                goods.paid_amount = total_paid_before + pay_now
                goods.balance_amount = gi_amt - goods.paid_amount


                goods.party_payment = payment_date
                goods.save(update_fields=["paid_amount", "balance_amount", "party_payment"])
                updated_goods.append(goods)
                paid_amount -= pay_now

            if updated_goods and original_paid_amount > 0:
                PaymentRecord.objects.create(
                    goods_info=updated_goods[0],  # store first for reference
                    consignment=goods.consignment,
                    party=party,
                    balance_amount=sum(g.balance_amount for g in updated_goods),
                    paid_amount=original_paid_amount,
                    payment_type="Part Payment",
                    payment_mode=payment_mode,
                    payment_date=payment_date,
                    remark=remark,
                    created_at=timezone.now(),
                )
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"error": "No valid goods found for payment."}, status=400)

        except Exception as e:
            traceback.print_exc()
            return JsonResponse({"error": f"Server error: {str(e)}"}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
#___________________________________________________________________

def statement_view(request):
    return render(request, "statement.html")

#___________________________________________________________________

# ──────────────────────────────────────────────────────────────
# PETROL MANAGEMENT VIEWS
# ──────────────────────────────────────────────────────────────

def petrol_pump_list(request):
    """Render the petrol pump list page."""
    return render(request, "petrol_pump.html")


@require_http_methods(["GET"])
def get_petrol_pumps(request):
    """
    Return all petrol pumps as JSON.
    Useful for dynamic frontend rendering via JavaScript.
    """
    petrolPumps = PetrolPump.objects.all().order_by("-date_added")
    formatted = [
        {
            "id": p.id,
            "petrol_pump_name": p.petrol_pump_name,
            "owner_name": p.owner_name,
            "owner_phone": p.owner_phone,
            "address": p.address,
            "date_added": p.date_added.strftime("%Y-%m-%d"),
        }
        for p in petrolPumps
    ]
    return JsonResponse({"success": True, "petrol_pumps": formatted})


@csrf_exempt
@require_http_methods(["POST"])
def add_petrol_pump(request):
    """
    Add a new petrol pump.
    Optional: owner_name, owner_phone, address
    Required: petrol_pump_name, date_added
    """
    try:
        data = json.loads(request.body)
        petrol_pump_name = data.get("petrol_pump_name")
        date_added = parse_date(data.get("date_added"))
        address = data.get("address")
        owner_name = data.get("owner_name") or None
        owner_phone = data.get("owner_phone") or None

        if not petrol_pump_name or not date_added:
            return JsonResponse({"success": False, "message": "Missing required fields."})

        petrolPump = PetrolPump.objects.create(
            petrol_pump_name=petrol_pump_name,
            owner_name=owner_name,
            owner_phone=owner_phone,
            address=address,
            date_added=date_added,
        )

        return JsonResponse({
            "success": True,
            "petrol_pump": {
                "id": petrolPump.id,
                "petrol_pump_name": petrolPump.petrol_pump_name,
                "owner_name": petrolPump.owner_name,
                "owner_phone": petrolPump.owner_phone,
                "address": petrolPump.address,
                "date_added": petrolPump.date_added.strftime("%Y-%m-%d"),
            }
        })

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def edit_petrol_pump(request, petrol_pump_id):
    """
    Edit an existing vehicle.
    Requires: petrol_pump_name, date_added, address
    Optional: owner_name, owner_phone
    """
    try:
        data = json.loads(request.body)
        petrol_pump_name = data.get("petrol_pump_name")
        date_added = parse_date(data.get("date_added"))
        address = data.get("address")
        owner_name = data.get("owner_name") or None
        owner_phone = data.get("owner_phone") or None

        if not petrol_pump_name or not date_added:
            return JsonResponse({"success": False, "message": "Missing required fields."})

        petrolPump = get_object_or_404(PetrolPump, id=petrol_pump_id)
        petrolPump.petrol_pump_name = petrol_pump_name
        petrolPump.date_added = date_added
        petrolPump.owner_name = owner_name
        petrolPump.owner_phone = owner_phone
        petrolPump.address = address
        petrolPump.save()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def delete_petrol_pump(request, petrol_pump_id):
    """
    Delete a petrol pump by ID using POST request.
    """
    try:
        petrolPump = get_object_or_404(PetrolPump, id=petrol_pump_id)
        petrolPump.delete()
        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    

# ──────────────────────────────────────────────────────────────
# STAFF/EMPLOYEE MANAGEMENT VIEWS
# ──────────────────────────────────────────────────────────────

def staffEmployee_list(request):
    """Render the staff/employee list page."""
    return render(request, "staff_employee.html")


@require_http_methods(["GET"])
def get_staff_employees(request):
    """
    Return all staff/employee  as JSON.
    Useful for dynamic frontend rendering via JavaScript.
    """
    staffEmployees = StaffEmployee.objects.all().order_by("-date_added")
    formatted = [
        {
            "id": s.id,
            "staff_employee_name": s.staff_employee_name,
            "phone_no": s.phone_no,
            "address": s.address,
            "date_added": s.date_added.strftime("%Y-%m-%d"),
        }
        for s in staffEmployees
    ]
    return JsonResponse({"success": True, "staff_employees": formatted})


@csrf_exempt
@require_http_methods(["POST"])
def add_staff_employee(request):
    """
    Add a new staff/employee.
    Optional: phone_no, address
    Required: staff_employee_name, date_added
    """
    try:
        data = json.loads(request.body)
        staff_employee_name = data.get("staff_employee_name")
        date_added = parse_date(data.get("date_added"))
        address = data.get("address")
        phone_no = data.get("phone_no") or None

        if not staff_employee_name or not date_added:
            return JsonResponse({"success": False, "message": "Missing required fields."})

        staffEmployee = StaffEmployee.objects.create(
            staff_employee_name=staff_employee_name,
            phone_no=phone_no,
            address=address,
            date_added=date_added,
        )

        return JsonResponse({
            "success": True,
            "petrol_pump": {
                "id": staffEmployee.id,
                "staff_employee_name": staffEmployee.staff_employee_name,
                "phone_no": staffEmployee.phone_no,
                "address": staffEmployee.address,
                "date_added": staffEmployee.date_added.strftime("%Y-%m-%d"),
            }
        })

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def edit_staff_employee(request, staff_employee_id):
    """
    Edit an existing vehicle.
    Requires: staff_employee_name, date_added, address
    Optional: phone_no
    """
    try:
        data = json.loads(request.body)
        staff_employee_name = data.get("staff_employee_name")
        date_added = parse_date(data.get("date_added"))
        address = data.get("address")
        phone_no = data.get("phone_no") or None

        if not staff_employee_name or not date_added:
            return JsonResponse({"success": False, "message": "Missing required fields."})

        staffEmployee = get_object_or_404(StaffEmployee, id=staff_employee_id)
        staffEmployee.staff_employee_name = staff_employee_name
        staffEmployee.date_added = date_added
        staffEmployee.phone_no = phone_no
        staffEmployee.address = address
        staffEmployee.save()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def delete_staff_employee(request, staff_employee_id):
    """
    Delete a staff/employee by ID using POST request.
    """
    try:
        staffEmployee = get_object_or_404(StaffEmployee, id=staff_employee_id)
        staffEmployee.delete()
        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    
# ──────────────────────────────────────────────────────────────
# VEHICLE  VIEW  
# ──────────────────────────────────────────────────────────────\
@csrf_exempt
def vehicle_view(request):
    vehicle_number = request.GET.get('vehicle_number')

    try:
        # 1. Get vehicle instance
        vehicle = Vehicle.objects.get(vehicle_number=vehicle_number)

        # 2. Get all consignments of this vehicle and sum the Balance_Amount
        from accounts.models import Consignment  # Adjust if needed

        total_balance = (
            Consignment.objects
            .filter(Vehicle_No=vehicle)
            .aggregate(total=models.Sum('Balance_Amount'))['total'] or 0
        )

        return JsonResponse({
            "success": True,
            "balance_amount": float(total_balance)
        })

    except Vehicle.DoesNotExist:
        return JsonResponse({
            "success": False,
            "error": "Vehicle not found"
        }, status=404)
    

@csrf_exempt
def record_vehicle_expense(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            vehicle_number = data.get("vehicle_number")
            amount = Decimal(data.get("amount", 0))
            payment_date = data.get("payment_date")
            payment_mode = data.get("payment_mode")
            remark = data.get("remark", "")

            vehicle = Vehicle.objects.get(vehicle_number=vehicle_number)

            # Get consignments with balance > 0 for this vehicle
            consignments = Consignment.objects.filter(
                Vehicle_No=vehicle, Balance_Amount__gt=0
            ).order_by('CNID')

            remaining = amount
            for consignment in consignments:
                if remaining <= 0:
                    break
                if consignment.Balance_Amount >= remaining:
                    consignment.Balance_Amount -= remaining
                    remaining = 0
                else:
                    remaining -= consignment.Balance_Amount
                    consignment.Balance_Amount = 0
                consignment.save()

            # Save record in RecordExpense
            RecordExpense.objects.create(
                expense_account_type="Vehicle",
                expense_account_name=vehicle_number,
                paid_amount=amount,
                payment_date=payment_date,
                payment_mode=payment_mode,
                remark=remark,
                created_at=timezone.now()
            )

            return JsonResponse({"success": True})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})


# ──────────────────────────────────────────────────────────────
# RECORD EXPENSE VIEW
# ──────────────────────────────────────────────────────────────

@require_http_methods(["GET", "POST"])
def record_expense_view(request):
    # Handle AJAX request for dropdown data
    if request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        pumps = PetrolPump.objects.all().order_by("petrol_pump_name")
        staff_employees = StaffEmployee.objects.all().order_by("staff_employee_name")
        vehicles = Vehicle.objects.all().order_by("vehicle_number")  # ✅ Get all vehicles

        pumps_data = [{"id": pump.id, "name": pump.petrol_pump_name} for pump in pumps]
        staff_data = [{"id": s.id, "name": s.staff_employee_name} for s in staff_employees]
        vehicle_data = [v.vehicle_number for v in vehicles]  # ✅ Convert to list of strings

        return JsonResponse({
            "success": True,
            "pumps": pumps_data,
            "staff_employees": staff_data,
            "vehicles": vehicle_data  # ✅ Include this in response
        })
    
    # Handle form submission
    elif request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            
            # Create new RecordExpense instance
            expense = RecordExpense(
                expense_account_name=data.get('expense_account_name'),
                expense_account_type=data.get('expense_account_type'),
                paid_amount=data.get('paid_amount'),
                payment_date=datetime.strptime(data.get('payment_date'), '%Y-%m-%d').date(),
                payment_mode=data.get('payment_mode'),
                remark=data.get('remark', '')
            )
            expense.save()
            
            return JsonResponse({
                "success": True,
                "message": "Expense recorded successfully!"
            })
            
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=400)
    
    # Regular GET request - render the template
    return render(request, "record_expense.html")



@require_http_methods(["GET", "POST"])
def staff_employees_view(request):
    # Handle AJAX request for dropdown data (GET)
    if request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        staff_employees = StaffEmployee.objects.all().order_by("staff_employee_name")

        data = {
            "success": True,
            "staff_employees": [{"id": s.id, "name": s.staff_employee_name} for s in staff_employees],
        }
        return JsonResponse(data)
    
    # Handle form submission (POST)
    elif request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            
            # Create new StaffEmployee instance
            staff_employee = StaffEmployee(
                staff_employee_name=data.get('staff_employee_name'),
                mobile_number=data.get('mobile_number', ''),
                address=data.get('address', ''),
                designation=data.get('designation', ''),
                salary=data.get('salary', 0),
                joining_date=datetime.strptime(data.get('joining_date'), '%Y-%m-%d').date() if data.get('joining_date') else None,
                # Add other fields as needed based on your StaffEmployee model
            )
            staff_employee.save()
            
            return JsonResponse({
                "success": True,
                "message": "Staff/Employee added successfully!",
                "staff_employee_id": staff_employee.id
            })
            
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=400)
    
    # Regular GET request - render the template
    elif request.method == 'GET':
        return render(request, "staff_employees.html")
    
    # Invalid request
    return JsonResponse({"success": False, "message": "Invalid request"}, status=400)



# ──────────────────────────────────────────────────────────────
# PARTY MAINTAINANCE VIEW
# ──────────────────────────────────────────────────────────────
def party_maintainance_view(request):
    selected_party_id = request.GET.get('party_id')
    selected_party = None
    debit_entries = []
    credit_entries = []
    total_debit = 0
    total_credit = 0
    balance = 0

    parties = Consignee.objects.all()

    if selected_party_id:
        selected_party = get_object_or_404(Consignee, id=selected_party_id)

        # ✅ Fix: use selected_party.id instead of the object
        # Get debit entries (consignments related to this party)
        debit_entries = Consignment.objects.filter(consignee_id=selected_party.id)
        print("Debit Entries:", debit_entries)

        total_debit = debit_entries.aggregate(total=Sum('total_fare'))['total'] or 0
        print("Total Debit:", total_debit)
        # Get credit entries (payment records related to this party)
        credit_entries = PaymentRecord.objects.filter(party=selected_party)
        total_credit = credit_entries.aggregate(total=Sum('paid_amount'))['total'] or 0

        balance = total_debit - total_credit

    return render(request, 'party_maintainance.html', {
        'parties': parties,
        'selected_party': selected_party,
        'debit_entries': debit_entries,
        'credit_entries': credit_entries,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'balance': balance,
    })


# def party_maintainance_view(request):
#     selected_party_id = request.GET.get('party_id')
#     selected_party = None
#     debit_entries = []
#     credit_entries = []
#     total_debit = 0
#     total_credit = 0
#     balance = 0

#     parties = Consignee.objects.all()

#     if selected_party_id:
#         selected_party = get_object_or_404(Consignee, id=selected_party_id)

#         # Get debit entries (consignments related to this party)
#         debit_entries = Consignment.objects.filter(consignee=selected_party)
#         total_debit = debit_entries.aggregate(total=Sum('total_fare'))['total'] or 0

#         # Get credit entries (payment records related to this party)
#         credit_entries = PaymentRecord.objects.filter(party=selected_party)
#         total_credit = credit_entries.aggregate(total=Sum('paid_amount'))['total'] or 0

#         # Compute balance
#         balance = total_debit - total_credit

#     return render(request, 'party_maintainance.html', {
#         'parties': parties,
#         'selected_party': selected_party,
#         'debit_entries': debit_entries,
#         'credit_entries': credit_entries,
#         'total_debit': total_debit,
#         'total_credit': total_credit,
#         'balance': balance,
#     })

# ──────────────────────────────────────────────────────────────
# VEHICLE MAINTAINANCE VIEW
# ──────────────────────────────────────────────────────────────
def vehicle_maintainance_view(request):
    selected_vehicle_id = request.GET.get('vehicle_id')
    vehicles = Vehicle.objects.all()

    if selected_vehicle_id:
        selected_vehicle = get_object_or_404(Vehicle, id=selected_vehicle_id)

        # Debit: Consignments where vehicle was used (Account Receivable)
        debit_entries = Consignment.objects.filter(Vehicle_No=selected_vehicle)

        # Credit: RecordExpense where account_type is 'Vehicle' and account_name = selected vehicle number
        credit_entries = RecordExpense.objects.filter(
            expense_account_type="Vehicle",
            expense_account_name=selected_vehicle.vehicle_number
        )

        # Use Balance_Amount as the debit value
        total_debit = debit_entries.aggregate(total=Sum('Balance_Amount'))['total'] or 0
        total_credit = credit_entries.aggregate(total=Sum('paid_amount'))['total'] or 0
        balance = total_debit - total_credit
    else:
        selected_vehicle = None
        debit_entries = credit_entries = []
        total_debit = total_credit = balance = 0

    return render(request, 'vehicle_maintainance.html', {
        'vehicles': vehicles,
        'selected_vehicle': selected_vehicle,
        'debit_entries': debit_entries,
        'credit_entries': credit_entries,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'balance': balance,
    })


# ──────────────────────────────────────────────────────────────
# Record Payment List View
# ──────────────────────────────────────────────────────────────
def record_payment_list_view(request):
    records = PaymentRecord.objects.all().order_by('-payment_date')
    return render(request, 'record_payment_list.html', {'payment_records': records})

@csrf_exempt
def delete_payment_record(request, pk):
    if request.method == "POST":
        try:
            record = get_object_or_404(PaymentRecord, pk=pk)

            # ✅ Restore payment in GoodsInfo (for both Parcha & Part Payment)
            if record.goods_info:   # Part Payment or Parcha stored with goods_info
                goods = record.goods_info
                goods.balance_amount = (goods.balance_amount or 0) + record.paid_amount
                goods.paid_amount = (goods.paid_amount or 0) - record.paid_amount
                goods.save(update_fields=["balance_amount", "paid_amount"])

            elif record.consignment:   # If only consignment is linked (Parcha case)
                goods_qs = GoodsInfo.objects.filter(consignment=record.consignment)

                # Restore proportionally or equally (depending on your logic)
                for goods in goods_qs:
                    # If this consignment only had one party → restore directly
                    goods.balance_amount = (goods.balance_amount or 0) + record.paid_amount
                    goods.paid_amount = (goods.paid_amount or 0) - record.paid_amount
                    goods.save(update_fields=["balance_amount", "paid_amount"])
                    break   # remove this if multiple goods need restoring

            # ✅ Now delete record
            record.delete()

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})



@csrf_exempt
def edit_payment_record(request, pk):
    if request.method == "POST":
        try:
            # Parse the JSON data from the request body
            data = json.loads(request.body)
            record = get_object_or_404(PaymentRecord, pk=pk)

            # Correct field names from model
            record.party = data.get('party_name', record.party)
            record.payment_date = data.get('payment_date', record.payment_date)
            record.payment_type = data.get('payment_type', record.payment_type)
            record.paid_amount = Decimal(data.get('paid_amount', record.paid_amount))
            record.payment_mode = data.get('payment_mode', record.payment_mode)

            record.save()

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# ──────────────────────────────────────────────────────────────
# Record Payment vehicle View
# ──────────────────────────────────────────────────────────────
def record_expense_list_view(request):
    # Get all types of expense records (Vehicle, Petrol Pump, Staff/Employee)
    records = RecordExpense.objects.all().order_by("-payment_date")
    return render(request, 'record_expense_list.html', {'expense_records': records})


@csrf_exempt
def delete_expense_record(request, pk):
    if request.method == "POST":
        try:
            record = get_object_or_404(RecordExpense, pk=pk)
            record.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
# ──────────────────────────────────────────────────────────────
# Cumulative  Report View
# ──────────────────────────────────────────────────────────────

def cumulative_report_view(request):
    goods = GoodsInfo.objects.select_related('consignment', 'to_party').all()
    payments = PaymentRecord.objects.select_related('consignment').all()

    party_entries = defaultdict(list)
    cumulative_totals = defaultdict(float)

    # 1. Add GOODS entries
    for item in goods:
        party_name = item.to_party.consignee_name if item.to_party else "Unknown"
        qty_10kg = item.quantity if item.unit == "10kg" else 0
        qty_20kg = item.quantity if item.unit == "20kg" else 0
        rate_10kg = item.rate if item.unit == "10kg" else 0
        rate_20kg = item.rate if item.unit == "20kg" else 0
        fare_10kg = qty_10kg * rate_10kg
        fare_20kg = qty_20kg * rate_20kg
        amount = fare_10kg + fare_20kg

        goods_entry = {
            'cn_no': item.consignment.CNID if item.consignment else '',
            'date': item.consignment.Loading_Date if item.consignment else item.created_at.date(),
            'truck_no': item.consignment.Vehicle_No.vehicle_number if item.consignment and item.consignment.Vehicle_No else '',
            'consignee': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
            'from_party': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
            'party': party_name,
            'qty_10kg': qty_10kg,
            'rate_10kg': rate_10kg,
            'qty_20kg': qty_20kg,
            'rate_20kg': rate_20kg,
            'fare_10kg': fare_10kg,
            'fare_20kg': fare_20kg,
            'total': amount,
            'amount': amount,
            'paid': 0,
            'is_payment': False,
            'sort_date': item.consignment.Loading_Date if item.consignment else item.created_at.date(),
            'created_at': item.created_at,
        }
        party_entries[party_name].append(goods_entry)

    # 2. Add PAYMENT entries
    for payment in payments:
        party_name = payment.party if payment.party else "Unknown"
        payment_entry = {
            'cn_no': '',
            'date': payment.payment_date,
            'truck_no': payment.consignment.Vehicle_No.vehicle_number if payment.consignment and payment.consignment.Vehicle_No else '',
            'consignee': '',
            'from_party': '',
            'party': party_name,
            'qty_10kg': 0,
            'rate_10kg': 0,
            'qty_20kg': 0,
            'rate_20kg': 0,
            'fare_10kg': 0,
            'fare_20kg': 0,
            'total': 0,
            'amount': 0,
            'paid': payment.paid_amount,
            'is_payment': True,
            'sort_date': payment.payment_date,
            'created_at': payment.created_at,
        }
        party_entries[party_name].append(payment_entry)

    # 3. Flatten and sort entries
    all_entries = list(chain.from_iterable(party_entries.values()))
    all_entries.sort(key=lambda x: (x['sort_date'], x['created_at']))

    # 4. Compute cumulative per party
    global_cumulative = defaultdict(float)

    for entry in all_entries:
        party = entry['party']
        if entry['is_payment']:
            global_cumulative[party] -= float(entry['paid'])
        else:
            global_cumulative[party] += float(entry['amount'])

        entry['cumulative'] = round(abs(global_cumulative[party]), 2)

    # 5. Compute totals (only for goods entries)
    total_10kg = 0
    total_rate_10kg = 0
    total_20kg = 0
    total_rate_20kg = 0
    total_fare_10kg = 0
    total_fare_20kg = 0

    for entry in all_entries:
        if not entry['is_payment']:
            total_10kg += entry['qty_10kg']
            total_rate_10kg += entry['rate_10kg']
            total_20kg += entry['qty_20kg']
            total_rate_20kg += entry['rate_20kg']
            total_fare_10kg += entry['fare_10kg']
            total_fare_20kg += entry['fare_20kg']  # Add this line

    return render(request, 'cumulative_report.html', {
        'entries': all_entries,
        'total_10kg': total_10kg,
        'total_rate_10kg': total_rate_10kg,
        'total_20kg': total_20kg,
        'total_rate_20kg': total_rate_20kg,
        'total_fare_10kg': total_fare_10kg,
        'total_fare_20kg': total_fare_20kg,  # Add this
    })



#my old code 
# def cumulative_report_view(request):
#     goods = GoodsInfo.objects.select_related('consignment', 'to_party').all()
#     payments = PaymentRecord.objects.select_related('consignment').all()

#     party_data = defaultdict(list)
#     cumulative_totals = defaultdict(float)
#     party_entries = defaultdict(list)

#     # 1. Add GOODS entries
#     for item in goods:
#         party_name = item.to_party.consignee_name if item.to_party else "Unknown"

#         qty_10kg = item.quantity if item.unit == "10kg" else 0
#         qty_20kg = item.quantity if item.unit == "20kg" else 0
#         rate_10kg = item.rate if item.unit == "10kg" else 0
#         rate_20kg = item.rate if item.unit == "20kg" else 0
#         fare_10kg = qty_10kg * rate_10kg
#         fare_20kg = qty_20kg * rate_20kg
#         amount = fare_10kg + fare_20kg

#         sort_date = item.created_at  # use datetime for accurate sorting

#         goods_entry = {
#             'cn_no': item.consignment.CNID if item.consignment else '',
#             'date': item.consignment.Booking_Date if item.consignment else item.created_at.date(),
#             'truck_no': item.consignment.Vehicle_No.vehicle_number if item.consignment and item.consignment.Vehicle_No else '',
#             'consignee': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
#             'party': party_name,
#             'qty_10kg': qty_10kg,
#             'rate_10kg': rate_10kg,
#             'qty_20kg': qty_20kg,
#             'rate_20kg': rate_20kg,
#             'fare_10kg': fare_10kg,
#             'fare_20kg': fare_20kg,
#             'total': amount,
#             'amount': amount,
#             'paid': 0,
#             'is_payment': False,
#             'sort_date': sort_date,
#             'created_at': item.created_at,
#         }
#         party_entries[party_name].append(goods_entry)

#     # 2. Add PAYMENT entries from PaymentRecord
#     for payment in payments:
#         party_name = payment.party or "Unknown"
#         sort_date = payment.created_at  # datetime (accurate)

#         payment_entry = {
#             'cn_no': '',
#             'date': payment.payment_date,  # for display only
#             'truck_no': payment.consignment.Vehicle_No.vehicle_number if payment.consignment and payment.consignment.Vehicle_No else '',
#             'consignee': '',
#             'party': party_name,
#             'qty_10kg': 0,
#             'rate_10kg': 0,
#             'qty_20kg': 0,
#             'rate_20kg': 0,
#             'fare_10kg': 0,
#             'fare_20kg': 0,
#             'total': 0,
#             'amount': 0,
#             'paid': payment.paid_amount,
#             'is_payment': True,
#             'sort_date': sort_date,
#             'created_at': payment.created_at,
#         }
#         party_entries[party_name].append(payment_entry)

#     # 3. Sort and calculate cumulative
#     for party_name in party_entries:
#         party_entries[party_name].sort(key=lambda x: (x['sort_date'], x['created_at']))
#         cumulative_totals[party_name] = 0.0

#         for entry in party_entries[party_name]:
#             if entry['is_payment']:
#                 cumulative_totals[party_name] -= float(entry['paid'])
#             else:
#                 cumulative_totals[party_name] += float(entry['amount'])

#             entry['cumulative'] = round(abs(cumulative_totals[party_name]), 2)

#         party_data[party_name] = party_entries[party_name]

#     # Sort all parties alphabetically
#     party_data = dict(sorted(party_data.items()))

#     return render(request, 'cumulative_report.html', {
#         'party_data': party_data
#     })

# This is shweta's code 
# def cumulative_report_view(request):
#     goods = GoodsInfo.objects.select_related('consignment', 'to_party').all()
#     payments = PaymentRecord.objects.select_related('consignment').all()

#     party_data = defaultdict(list)
#     cumulative_totals = defaultdict(float)
#     party_entries = defaultdict(list)

#     # 1. Add GOODS entries
#     for item in goods:
#         party_name = item.to_party.consignee_name if item.to_party else "Unknown"

#         qty_10kg = item.quantity if item.unit == "10kg" else 0
#         qty_20kg = item.quantity if item.unit == "20kg" else 0
#         rate_10kg = item.rate if item.unit == "10kg" else 0
#         rate_20kg = item.rate if item.unit == "20kg" else 0
#         fare_10kg = qty_10kg * rate_10kg
#         fare_20kg = qty_20kg * rate_20kg
#         amount = fare_10kg + fare_20kg

#         sort_date = item.created_at  # use datetime for accurate sorting

#         goods_entry = {
#             'cn_no': item.consignment.CNID if item.consignment else '',
#             'date': item.consignment.Booking_Date if item.consignment else item.created_at.date(),
#             'truck_no': item.consignment.Vehicle_No.vehicle_number if item.consignment and item.consignment.Vehicle_No else '',
#             'consignee': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
#             'party': party_name,
#             'qty_10kg': qty_10kg,
#             'rate_10kg': rate_10kg,
#             'qty_20kg': qty_20kg,
#             'rate_20kg': rate_20kg,
#             'fare_10kg': fare_10kg,
#             'fare_20kg': fare_20kg,
#             'total': amount,
#             'amount': amount,
#             'paid': 0,
#             'is_payment': False,
#             'sort_date': sort_date,
#             'created_at': item.created_at,
#         }
#         party_entries[party_name].append(goods_entry)

#     # 2. Add PAYMENT entries from PaymentRecord
#     for payment in payments:
#         party_name = payment.party or "Unknown"
#         sort_date = payment.created_at  # datetime (accurate)

#         payment_entry = {
#             'cn_no': '',
#             'date': payment.payment_date,  # for display only
#             #'truck_no': payment.consignment.Vehicle_No.vehicle_number if payment.consignment and payment.consignment.Vehicle_No else '',
#             'consignee': '',
#             'party': party_name,
#             'qty_10kg': 0,
#             'rate_10kg': 0,
#             'qty_20kg': 0,
#             'rate_20kg': 0,
#             'fare_10kg': 0,
#             'fare_20kg': 0,
#             'total': 0,
#             'amount': 0,
#             'paid': payment.paid_amount,
#             'is_payment': True,
#             'sort_date': sort_date,
#             'created_at': payment.created_at,
#         }
#         party_entries[party_name].append(payment_entry)

#     # 3. Sort and calculate cumulative
#     for party_name in party_entries:
#         party_entries[party_name].sort(key=lambda x: (x['sort_date'], x['created_at']))
#         cumulative_totals[party_name] = 0.0

#         for entry in party_entries[party_name]:
#             if entry['is_payment']:
#                 cumulative_totals[party_name] -= float(entry['paid'])
#             else:
#                 cumulative_totals[party_name] += float(entry['amount'])

#             entry['cumulative'] = round(abs(cumulative_totals[party_name]), 2)

#         party_data[party_name] = party_entries[party_name]

#     # Sort all parties alphabetically
#     party_data = dict(sorted(party_data.items()))

#     return render(request, 'cumulative_report.html', {
#         'party_data': party_data
#     })
# def cumulative_report_view(request):
#     goods = GoodsInfo.objects.select_related(
#         'consignment', 'to_party', 'consignment__Vehicle_No', 'consignment__consignee'
#     ).order_by('created_at')

#     party_data = defaultdict(list)
#     party_entries = defaultdict(list)

#     # 1. GOODS ENTRIES
#     for item in goods:
#         party_name = item.to_party.consignee_name if item.to_party else "Unknown"

#         if item.quantity:
#             qty_10kg = item.quantity if item.unit == "10kg" else 0
#             qty_20kg = item.quantity if item.unit == "20kg" else 0
#             rate_10kg = item.rate if item.unit == "10kg" else 0
#             rate_20kg = item.rate if item.unit == "20kg" else 0
#             fare_10kg = qty_10kg * rate_10kg
#             fare_20kg = qty_20kg * rate_20kg
#             total_amount = item.gi_amount if item.gi_amount else (fare_10kg + fare_20kg)

#             sort_date = item.consignment.Booking_Date if item.consignment and item.consignment.Booking_Date else item.created_at
#             if isinstance(sort_date, date) and not isinstance(sort_date, datetime):
#                 sort_date = datetime.combine(sort_date, datetime.min.time())
#             if not timezone.is_aware(sort_date):
#                 sort_date = timezone.make_aware(sort_date)

#             entry = {
#                 'cn_no': item.consignment.CNID if item.consignment else '',
#                 'date': sort_date,
#                 'truck_no': item.consignment.Vehicle_No.vehicle_number if item.consignment and item.consignment.Vehicle_No else '',
#                 'consignee': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
#                 'party': party_name,
#                 'qty_10kg': qty_10kg,
#                 'rate_10kg': rate_10kg,
#                 'qty_20kg': qty_20kg,
#                 'rate_20kg': rate_20kg,
#                 'fare_10kg': fare_10kg,
#                 'fare_20kg': fare_20kg,
#                 'total': total_amount,
#                 'amount': total_amount,
#                 'paid': 0,
#                 'is_payment': False,
#                 'sort_date': sort_date,
#                 'created_at': item.created_at,
#             }
#             party_entries[party_name].append(entry)

#     # 2. PAYMENT ENTRIES (from PaymentRecord table)
#     payments = PaymentRecord.objects.all().order_by("payment_date")
#     for payment in payments:
#         party_name = payment.party or "Unknown"
#         payment_date = payment.payment_date or payment.created_at

#         if isinstance(payment_date, date) and not isinstance(payment_date, datetime):
#             payment_date = datetime.combine(payment_date, datetime.min.time())
#         if not timezone.is_aware(payment_date):
#             payment_date = timezone.make_aware(payment_date)

#         payment_entry = {
#             'cn_no': '',
#             'date': payment_date,
#             'truck_no': '',
#             'consignee': '',
#             'party': party_name,
#             'qty_10kg': 0,
#             'rate_10kg': 0,
#             'qty_20kg': 0,
#             'rate_20kg': 0,
#             'fare_10kg': 0,
#             'fare_20kg': 0,
#             'total': 0,
#             'amount': 0,
#             'paid': float(payment.paid_amount),
#             'is_payment': True,
#             'sort_date': payment_date,
#             'created_at': payment.created_at,
#         }
#         party_entries[party_name].append(payment_entry)

#     # 3. FINAL SORTING AND CUMULATIVE
#     for party in party_entries:
#         party_entries[party].sort(key=lambda x: (x['sort_date'], not x['is_payment'], x['created_at']))

#         cumulative = 0.0
#         for entry in party_entries[party]:
#             if entry['is_payment']:
#                 cumulative -= entry['paid']
#             else:
#                 cumulative += float(entry['amount'])
#             entry['cumulative'] = round(abs(cumulative), 2)

#         party_data[party] = party_entries[party]

#     return render(request, 'cumulative_report.html', {
#         'party_data': dict(sorted(party_data.items()))
#     })

# from openpyxl.drawing.image import Image as XLImage

# ──────────────────────────────────────────────────────────────
# Export Cumulative Report in Excel and PDF formats
# ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def export_cumulative_excel(request):
    global_search = request.GET.get('global_search', '').strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cumulative Report"

    # --- Header: Company Details and Logo ---
    ws.merge_cells('A1:N1')
    ws.merge_cells('A2:N2')
    ws.merge_cells('A3:N3')
    ws.merge_cells('A4:N4')

    ws['A1'] = "DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS"
    ws['A2'] = "IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING"
    ws['A3'] = "18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304"
    ws['A4'] = "Mob: +91 9830357491 | Email: kakashri7314@gmail.com"

    for i in range(1, 5):
        cell = ws[f"A{i}"]
        cell.font = Font(bold=True if i == 1 else False, size=11)
        cell.alignment = Alignment(horizontal='center')

    # --- Logo ---
    try:
        from django.conf import settings
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'staticfiles', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'accounts', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
        ]
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 100
                img.height = 75
                ws.add_image(img, 'N1')
                break
    except Exception:
        pass

    ws.append([])  # Spacer row

    # --- Title ---
    ws.merge_cells('A6:N6')
    ws['A6'] = "Party Report"
    ws['A6'].font = Font(bold=True, size=14, color="6A1B9A")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.append([])

    # --- Table Headers ---
    headers = [
        'CN No', 'Date', 'Truck No', 'Consignee','From Party', 'Party',
        '10 KG Qty', '10 KG Rate', '20 KG Qty', '20 KG Rate',
        '10 KG Fare', '20 KG Fare', 'Total',
        'Paid', 'Balance'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='E5CCFF', end_color='E5CCFF', fill_type='solid')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='000000')
        cell.alignment = Alignment(horizontal='center')
        # cell.fill = header_fill

    # --- Data Fetching ---
    goods = GoodsInfo.objects.select_related('consignment', 'to_party').all()
    payments = PaymentRecord.objects.select_related('consignment').all()
    party_entries = defaultdict(list)

    for item in goods:
        party = item.to_party.consignee_name if item.to_party else "Unknown"
        if global_search and global_search.lower() not in party.lower():
            continue

        qty_10 = item.quantity if item.unit == "10kg" else 0
        qty_20 = item.quantity if item.unit == "20kg" else 0
        rate_10 = item.rate if item.unit == "10kg" else 0
        rate_20 = item.rate if item.unit == "20kg" else 0
        fare_10 = qty_10 * rate_10
        fare_20 = qty_20 * rate_20
        total = fare_10 + fare_20

        party_entries[party].append({
            'cn_no': item.consignment.CNID if item.consignment else '',
            'date': item.consignment.Booking_Date if item.consignment else item.created_at.date(),
            'truck_no': item.consignment.Vehicle_No.vehicle_number if item.consignment and item.consignment.Vehicle_No else '',
            'consignee': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
            'from_party': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
            'party': party,
            'qty_10kg': qty_10,
            'rate_10kg': rate_10,
            'qty_20kg': qty_20,
            'rate_20kg': rate_20,
            'fare_10kg': fare_10,
            'fare_20kg': fare_20,
            'total': total,
            'paid': 0,
            'is_payment': False,
            'sort_date': item.consignment.Booking_Date if item.consignment else item.created_at.date(),
            'created_at': item.created_at
        })

    for p in payments:
        party = p.party or "Unknown"
        if global_search and global_search.lower() not in party.lower():
            continue

        party_entries[party].append({
            'cn_no': 'PAYMENT',
            'date': p.payment_date,
            'truck_no': p.consignment.Vehicle_No.vehicle_number if p.consignment and p.consignment.Vehicle_No else '',
            'consignee': '',
            'from_party': '',
            'party': party,
            'qty_10kg': 0,
            'rate_10kg': 0,
            'qty_20kg': 0,
            'rate_20kg': 0,
            'fare_10kg': 0,
            'fare_20kg': 0,
            'total': 0,
            'paid': p.paid_amount,
            'is_payment': True,
            'sort_date': p.payment_date,
            'created_at': p.created_at
        })

    # --- Totals Calculation (exclude 20 KG Rate) ---
    total_10kg = total_rate_10kg = total_20kg = 0
    total_fare_10kg = total_fare_20kg = 0
    overall_total = overall_paid = overall_balance = 0

    for party in party_entries:
        cumulative_balance = 0
        for entry in sorted(party_entries[party], key=lambda x: (x['sort_date'], x['created_at'])):
            if not entry['is_payment']:
                total_10kg += entry['qty_10kg']
                total_rate_10kg += entry['rate_10kg']
                total_20kg += entry['qty_20kg']
                total_fare_10kg += entry['fare_10kg']
                total_fare_20kg += entry['fare_20kg']
                overall_total += entry['total']
                cumulative_balance += entry['total']
            else:
                overall_paid += entry['paid']
                cumulative_balance -= entry['paid']
        overall_balance += cumulative_balance

    # --- Totals Row (skip 20 KG Rate) ---
    ws.append([
        "TOTALS", "", "", "", "",
        total_10kg,
        total_rate_10kg,
        total_20kg,
        "",  # 20 KG Rate not included
        round(total_fare_10kg, 2),
        round(total_fare_20kg, 2),
        round(overall_total, 2),
        round(overall_paid, 2),
        round(overall_balance, 2)
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FF0000')
        cell.alignment = Alignment(horizontal='center')

    # --- Data Rows ---
    for party in sorted(party_entries.keys()):
        party_entries[party].sort(key=lambda x: (x['sort_date'], x['created_at']))
        cumulative_balance = 0.0

        for row in party_entries[party]:
            if row['is_payment']:
                cumulative_balance -= float(row['paid'])
            else:
                cumulative_balance += float(row['total'])

            ws.append([
                row['cn_no'],
                row['date'].strftime('%d-%m-%Y') if row['date'] else '',
                row['truck_no'],
                row['consignee'],
                row['from_party'],  # Add this line
                row['party'],
                row['qty_10kg'],
                row['rate_10kg'],
                row['qty_20kg'],
                row['rate_20kg'],
                round(row['fare_10kg'], 2),
                round(row['fare_20kg'], 2),
                round(row['total'], 2),
                round(row['paid'], 2) if row['is_payment'] else 0,
                round(abs(cumulative_balance), 2)
            ])

    # --- Auto column width ---
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 25)

    # --- Return Excel File ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="cumulative_report.xlsx"'
    wb.save(response)
    return response



# ──────────────────────────────────────────────────────────────
# Export Cumulative Report in PDF format
# ──────────────────────────────────────────────────────────────
from reportlab.lib import colors  # ✅ Corrected import

@login_required(login_url="login")
def export_cumulative_pdf(request):
    global_search = request.GET.get('global_search', '').strip()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cumulative_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            leftMargin=1 * cm, rightMargin=1 * cm,
                            topMargin=1.5 * cm, bottomMargin=1 * cm)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    wrap_style = styles['Normal'].clone('WrapStyle')
    wrap_style.fontSize = 8  # slightly increased font
    wrap_style.leading = 9
    wrap_style.alignment = TA_CENTER

    # Header styles
    company_style = styles['Title'].clone('CompanyStyle')
    company_style.fontSize = 16
    company_style.textColor = colors.HexColor('#1a365d')
    company_style.alignment = TA_CENTER

    subtitle_style = styles['Normal'].clone('SubtitleStyle')
    subtitle_style.fontSize = 10
    subtitle_style.alignment = TA_CENTER

    address_style = styles['Normal'].clone('AddressStyle')
    address_style.fontSize = 9
    address_style.alignment = TA_CENTER

    # Logo
    try:
        from django.conf import settings
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'staticfiles', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        ]
        logo_cell = None
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=3 * cm, height=2.5 * cm)
                logo.hAlign = 'RIGHT'
                logo_cell = logo
                break
        else:
            logo_cell = Paragraph("DHANLAXMI", company_style)
    except Exception:
        logo_cell = Paragraph("DHANLAXMI", company_style)

    address_block = [
        Paragraph("<b>DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS</b>", company_style),
        Paragraph("<b>IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING</b>", subtitle_style),
        Paragraph("18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304", address_style),
        Paragraph("Mob: +91 9830357491 | Email: kakashri7314@gmail.com", address_style),
    ]

    header_table = Table(
        [[address_block, logo_cell]],
        colWidths=[16 * cm, 5 * cm],
        hAlign='CENTER'
    )
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),  # logo beside address
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 10))

    # Title
    title_style = styles['Title'].clone('TitleStyle')
    title_style.fontSize = 13
    title_style.textColor = colors.HexColor('#6a1b9a')
    title_style.alignment = TA_CENTER
    elements.append(Paragraph("<b>Party Report</b>", title_style))
    elements.append(Spacer(1, 8))

    # Table headers
    headers = [
        "S.No", "CN No", "Date", "Truck", "Consignee", "From Party", "Party",
        "10KG", "10KG Rate", "20KG", "20KG Rate", "10KG Fare", "20KG Fare", "Total", "Paid", "Balance"
    ]
    header_row = [Paragraph(f"<b>{header}</b>", wrap_style) for header in headers]
    data = [header_row]

    # Function to wrap text
    def create_wrapped_paragraph(text, max_chars=10):
        if not text:
            return Paragraph("", wrap_style)
        text = str(text)
        if len(text) <= max_chars:
            return Paragraph(text, wrap_style)
        wrapped_text = ""
        words = text.split()
        current_line = ""
        for word in words:
            test_line = current_line + " " + word if current_line else word
            if len(test_line) <= max_chars:
                current_line = test_line
            else:
                if current_line:
                    wrapped_text += current_line + "<br/>"
                    current_line = word
                else:
                    wrapped_text += word[:max_chars] + "<br/>"
                    current_line = word[max_chars:]
        if current_line:
            wrapped_text += current_line
        return Paragraph(wrapped_text, wrap_style)

    goods = GoodsInfo.objects.select_related('consignment', 'to_party').all()
    payments = PaymentRecord.objects.select_related('consignment').all()
    party_entries = defaultdict(list)

    for item in goods:
        party = item.to_party.consignee_name if item.to_party else "Unknown"
        if global_search and global_search.lower() not in party.lower():
            continue
        qty_10 = item.quantity if item.unit == "10kg" else 0
        qty_20 = item.quantity if item.unit == "20kg" else 0
        rate_10 = item.rate if item.unit == "10kg" else 0
        rate_20 = item.rate if item.unit == "20kg" else 0
        fare_10 = qty_10 * rate_10
        fare_20 = qty_20 * rate_20
        total = fare_10 + fare_20

        party_entries[party].append({
            'cn_no': item.consignment.CNID if item.consignment else '',
            'date': item.consignment.Booking_Date if item.consignment else item.created_at.date(),
            'truck_no': item.consignment.Vehicle_No.vehicle_number if item.consignment and item.consignment.Vehicle_No else '',
            'consignee': item.consignment.consignee.consignee_name if item.consignment and item.consignment.consignee else '',
            'from_party': item.consignment.consigner.consigner_name if item.consignment and item.consignment.consigner else '',
            'party': party,
            'qty_10kg': qty_10,
            'rate_10kg': rate_10,
            'qty_20kg': qty_20,
            'rate_20kg': rate_20,
            'fare_10kg': fare_10,
            'fare_20kg': fare_20,
            'total': total,
            'paid': 0,
            'is_payment': False,
            'sort_date': item.consignment.Booking_Date if item.consignment else item.created_at.date(),
            'created_at': item.created_at,
        })

    for p in payments:
        party = p.party or "Unknown"
        if global_search and global_search.lower() not in party.lower():
            continue
        party_entries[party].append({
            'cn_no': "PAYMENT",
            'date': p.payment_date,
            'truck_no': p.consignment.Vehicle_No.vehicle_number if p.consignment and p.consignment.Vehicle_No else '',
            'consignee': '',
            'from_party': '',
            'party': party,
            'qty_10kg': 0,
            'rate_10kg': 0,
            'qty_20kg': 0,
            'rate_20kg': 0,
            'fare_10kg': 0,
            'fare_20kg': 0,
            'total': 0,
            'paid': p.paid_amount,
            'is_payment': True,
            'sort_date': p.payment_date,
            'created_at': p.created_at,
        })

    # Totals calculation
    total_10kg = total_rate_10kg = total_20kg = 0
    total_fare_10kg = total_fare_20kg = 0
    for party in party_entries:
        for entry in party_entries[party]:
            if not entry['is_payment']:
                total_10kg += entry['qty_10kg']
                total_rate_10kg += entry['rate_10kg']
                total_20kg += entry['qty_20kg']
                total_fare_10kg += entry['fare_10kg']
                total_fare_20kg += entry['fare_20kg']

    total_total = total_fare_10kg + total_fare_20kg
    total_paid = sum(p.paid_amount for p in payments if not global_search or global_search.lower() in (p.party or '').lower())
    total_balance = total_total - total_paid

    # Summary row
    summary_row = [
        Paragraph("<b>TOTALS</b>", wrap_style),
        Paragraph("", wrap_style), Paragraph("", wrap_style), Paragraph("", wrap_style),
        Paragraph("", wrap_style), Paragraph("", wrap_style), Paragraph("", wrap_style),
        Paragraph(f"<b>{int(total_10kg)}</b>", wrap_style),
        Paragraph(f"<b>{int(total_rate_10kg)}</b>", wrap_style),
        Paragraph(f"<b>{int(total_20kg)}</b>", wrap_style),
        Paragraph("", wrap_style),
        Paragraph(f"<b>{int(total_fare_10kg)}</b>", wrap_style),
        Paragraph(f"<b>{int(total_fare_20kg)}</b>", wrap_style),
        Paragraph(f"<b>{int(total_total)}</b>", wrap_style),
        Paragraph(f"<b>{int(total_paid)}</b>", wrap_style),
        Paragraph(f"<b>{int(total_balance)}</b>", wrap_style),
    ]
    data.append(summary_row)

    # Add data rows
    serial = 1
    for party in sorted(party_entries.keys()):
        party_entries[party].sort(key=lambda x: (x['sort_date'], x['created_at']))
        cumulative_balance = 0.0
        for row in party_entries[party]:
            if row['is_payment']:
                cumulative_balance -= float(row['paid'])
            else:
                cumulative_balance += float(row['total'])

            data_row = [
                Paragraph(str(serial), wrap_style),
                Paragraph(str(row['cn_no']), wrap_style),
                Paragraph(row['date'].strftime('%d-%m-%Y') if row['date'] else '', wrap_style),
                create_wrapped_paragraph(row['truck_no'], 8),
                create_wrapped_paragraph(row['consignee'], 10),
                create_wrapped_paragraph(row['from_party'], 8),
                create_wrapped_paragraph(row['party'], 10),
                Paragraph(str(int(row['qty_10kg'])), wrap_style),
                Paragraph(str(int(row['rate_10kg'])), wrap_style),
                Paragraph(str(int(row['qty_20kg'])), wrap_style),
                Paragraph(str(int(row['rate_20kg'])), wrap_style),
                Paragraph(str(int(row['fare_10kg'])), wrap_style),
                Paragraph(str(int(row['fare_20kg'])), wrap_style),
                Paragraph(str(int(row['total'])), wrap_style),
                Paragraph(str(int(row['paid']) if row['is_payment'] else 0), wrap_style),
                Paragraph(str(int(abs(cumulative_balance))), wrap_style),
            ]
            data.append(data_row)
            serial += 1

    col_widths = [
        1.0 * cm, 1.4 * cm, 1.6 * cm, 1.8 * cm, 2.2 * cm, 2.0 * cm, 2.2 * cm,
        1.2 * cm, 1.2 * cm, 1.2 * cm, 1.2 * cm, 1.4 * cm, 1.4 * cm, 1.4 * cm, 1.2 * cm, 1.4 * cm
    ]

    table = Table(data, repeatRows=1, colWidths=col_widths, hAlign='CENTER')
    table.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("WORDWRAP", (0, 0), (-1, -1), True),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 15))
    footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    footer_para = Paragraph(footer_text, address_style)
    elements.append(footer_para)

    doc.build(elements)
    return response


# ──────────────────────────────────────────────────────────────
# Vehicle Cumulative Report View
# ──────────────────────────────────────────────────────────────
from django.conf import settings
from datetime import datetime, time
@login_required(login_url='login')
def vehicle_cumulative_view(request):
    consignments = Consignment.objects.select_related('Vehicle_No', 'consignee').order_by('Loading_Date')
    vehicle_payments = RecordExpense.objects.filter(expense_account_type="Vehicle").order_by('payment_date')

    vehicle_entries = defaultdict(list)

    # ─── 1. Consignment Entries ─────────────────────────────
    for entry in consignments:
        vehicle_no = entry.Vehicle_No.vehicle_number if entry.Vehicle_No else "Unknown"
        to_pay = (
            float(entry.Truck_Freight or 0) +
            float(entry.Innam or 0) +
            float(entry.Extra_TF or 0) -
            float(entry.Advance_Amount or 0)
        )

        sort_date = timezone.make_aware(datetime.combine(entry.Loading_Date, time.min))

        vehicle_entry = {
            'is_payment': False,
            'date': entry.Loading_Date,
            'vehicle_no': vehicle_no,
            'cn_no': entry.Cn_No,
            'consignee': entry.consignee.consignee_name if entry.consignee else "",
            'truck_fare': float(entry.Truck_Freight or 0),
            'innam': float(entry.Innam or 0),
            'extra_tf': float(entry.Extra_TF or 0),
            'advance': float(entry.Advance_Amount or 0),
            'to_pay': to_pay,
            'amount': to_pay,
            'paid': 0,
            'remark': '',
            'sort_date': sort_date,
            'created_at': sort_date,
        }

        vehicle_entries[vehicle_no].append(vehicle_entry)

    # ─── 2. Payment Entries ────────────────────────────────
    for payment in vehicle_payments:
        vehicle_no = payment.expense_account_name or "Unknown"
        payment_dt = timezone.make_aware(datetime.combine(payment.payment_date, time.min))

        payment_entry = {
            'is_payment': True,
            'date': payment.payment_date,
            'vehicle_no': vehicle_no,
            'cn_no': 'PAYMENT',
            'consignee': '',
            'truck_fare': 0,
            'innam': 0,
            'extra_tf': 0,
            'advance': 0,
            'to_pay': 0,
            'amount': 0,
            'paid': float(payment.paid_amount or 0),
            'remark': payment.remark or '',
            'sort_date': payment_dt,
            'created_at': payment.created_at if timezone.is_aware(payment.created_at) else timezone.make_aware(payment.created_at),
        }

        vehicle_entries[vehicle_no].append(payment_entry)

    # ─── 3. Combine all entries and sort by date (latest first) ──────
    all_entries = []
    for vehicle_no, entries in vehicle_entries.items():
        # Combine all entries for this vehicle
        all_entries.extend(entries)

    # Sort all entries by date in descending order (latest first)
    # Primary sort by date, secondary sort by created_at for same-date entries
    all_entries.sort(key=lambda e: (e['sort_date'], e['created_at']), reverse=True)

    # ─── 4. Calculate Cumulative (but we need to calculate in chronological order) ───────────────────────────
    # For cumulative calculation, we need to process in chronological order (oldest first)
    # So we create a copy sorted in ascending order for calculation
    chronological_entries = sorted(all_entries, key=lambda e: (e['sort_date'], e['created_at']))
    
    cumulative_dict = defaultdict(float)
    for entry in chronological_entries:
        vehicle = entry['vehicle_no']

        if entry['is_payment']:
            cumulative_dict[vehicle] -= entry['paid']
        else:
            cumulative_dict[vehicle] += entry['amount']

        entry['cumulative'] = round(cumulative_dict[vehicle], 2)

    # ─── 5. Render to Template (entries are already sorted latest first) ─────────────────────────────
    return render(request, 'vehicle_cumulative.html', {
        'entries': all_entries,  # This maintains the latest-first order
        'global_search': request.GET.get('global_search', '')
    })

# ──────────────────────────────────────────────────────────────
# Vehicle Cumulative Report Excel Export
# ──────────────────────────────────────────────────────────────
@login_required(login_url='login')
def export_cumulative_vehicle_excel(request):
    global_search = request.GET.get("global_search", "").strip()

    entries = Consignment.objects.select_related('Vehicle_No', 'consignee').order_by('-Loading_Date')
    if global_search:
        entries = entries.filter(Vehicle_No__vehicle_number__icontains=global_search)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cumulative Report"

    # --- Header: Company Name and Info ---
    ws.merge_cells('A1:I1')
    ws.merge_cells('A2:I2')
    ws.merge_cells('A3:I3')
    ws.merge_cells('A4:I4')

    ws['A1'] = "DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS"
    ws['A2'] = "IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING"
    ws['A3'] = "18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304"
    ws['A4'] = "Mob: +91 9830357491 | Email: kakashri7314@gmail.com"

    for i in range(1, 5):
        ws[f"A{i}"].font = Font(bold=True if i == 1 else False, size=11)
        ws[f"A{i}"].alignment = Alignment(horizontal='center')

    # --- Add Logo if available ---
    try:
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'logo.png'),
        ]
        for path in logo_paths:
            if os.path.exists(path):
                img = XLImage(path)
                img.width = 100
                img.height = 75
                ws.add_image(img, 'I1')
                break
    except Exception:
        pass

    ws.append([])

    # --- Title Row ---
    ws.merge_cells('A6:I6')
    ws['A6'] = "Vehicle Cumulative Report"
    ws['A6'].font = Font(bold=True, size=14, color="6A1B9A")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.append([])

    # --- Table Headers ---
    headers = ['Bill No', 'Date', 'Truck No', 'Consignee', 'Truck Fare', 'Innam', 'Extra TF', 'Advance', 'Cumulative To Pay']
    ws.append(headers)
    header_fill = PatternFill(start_color='E5CCFF', end_color='E5CCFF', fill_type='solid')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # --- Table Data ---
    for entry in entries:
        truck_fare = float(entry.Truck_Freight or 0)
        innam = float(entry.Innam or 0)
        extra_tf = float(entry.Extra_TF or 0)
        advance = float(entry.Advance_Amount or 0)
        balance = float(entry.Balance_Amount or 0)

        ws.append([
            entry.Cn_No,
            entry.Loading_Date.strftime('%d-%m-%Y') if entry.Loading_Date else '',
            entry.Vehicle_No.vehicle_number if entry.Vehicle_No else '',
            entry.consignee.consignee_name if entry.consignee else '',
            truck_fare,
            innam,
            extra_tf,
            advance,
            balance,
        ])

    # --- Auto-fit Columns ---
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 25)

    # --- Return Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vehicle_cumulative_report.xlsx'
    wb.save(response)
    return response


# ──────────────────────────────────────────────────────────────
# Vehicle Cumulative Report PDF Export
# ──────────────────────────────────────────────────────────────
@login_required(login_url="login")
def export_cumulative_vehicle_pdf(request):
    
    global_search = request.GET.get('global_search', '').strip()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="vehicle_cumulative_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            leftMargin=1 * cm, rightMargin=1 * cm,
                            topMargin=1.5 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = []

    # Header styles
    company_style = styles['Title'].clone('CompanyStyle')
    company_style.fontSize = 16
    company_style.textColor = colors.HexColor('#1a365d')
    company_style.alignment = TA_CENTER

    subtitle_style = styles['Normal'].clone('SubtitleStyle')
    subtitle_style.fontSize = 10
    subtitle_style.alignment = TA_CENTER

    address_style = styles['Normal'].clone('AddressStyle')
    address_style.fontSize = 9
    address_style.alignment = TA_CENTER

    # Logo
    try:
        from django.conf import settings
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'accounts', 'static', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'staticfiles', 'images', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        ]
        logo_cell = None
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=3 * cm, height=2.5 * cm)
                logo.hAlign = 'RIGHT'
                logo_cell = logo
                break
        else:
            logo_cell = Paragraph("DHANLAXMI", company_style)
    except Exception:
        logo_cell = Paragraph("DHANLAXMI", company_style)

    address_block = [
        Paragraph("<b>DHANLAXMI AGRO ACTIVITIES & TRANSLOGISTICS</b>", company_style),
        Paragraph("<b>IMPORT | EXPORT | FRUIT PROCESSING AGRO MARKETING & TRADING</b>", subtitle_style),
        Paragraph("18, 3rd Floor, Dharmaveer Raje Sambhaji Vyapari Sankul, Ugaon, Tal-Niphad, Dist-Nashik - 422304", address_style),
        Paragraph("Mob: +91 9830357491 | Email: kakashri7314@gmail.com", address_style),
    ]

    header_table = Table([[address_block, logo_cell]], colWidths=[16 * cm, 5 * cm], hAlign='CENTER')
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))

    # Title
    title_style = styles['Title'].clone('TitleStyle')
    title_style.fontSize = 14
    title_style.textColor = colors.HexColor('#6a1b9a')
    title_style.alignment = TA_CENTER
    elements.append(Paragraph("<b>Vehicle Cumulative Report</b>", title_style))
    elements.append(Spacer(1, 12))

    # Table headers
    headers = [
        "S.No", "Bill No", "Date", "Truck No", "Consignee",
        "Truck Fare", "Innam", "Extra TF", "Advance", "Cumulative To Pay"
    ]
    data = [headers]

    entries = Consignment.objects.select_related('Vehicle_No', 'consignee').order_by('-Loading_Date')
    if global_search:
        entries = entries.filter(consignee__consignee_name__icontains=global_search)

    serial = 1
    for entry in entries:
        row = [
            serial,
            entry.Cn_No,
            entry.Loading_Date.strftime('%d-%m-%Y') if entry.Loading_Date else '',
            entry.Vehicle_No.vehicle_number if entry.Vehicle_No else '',
            entry.consignee.consignee_name if entry.consignee else '',
            f"{entry.Truck_Freight:.0f}",
            f"{entry.Innam:.0f}",
            f"{entry.Extra_TF:.0f}",
            f"{entry.Advance_Amount:.0f}",
            f"{entry.Balance_Amount:.0f}",
        ]
        data.append(row)
        serial += 1

    # Table rendering with styles
    col_widths = [1.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header row: bold and black
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        # All data rows: normal weight and black
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),  # Regular font for data

        # Optional light background for second row (can remove if not needed)
        # ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f0f8ff")),

        # Other styles
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 20))
    footer = Paragraph(f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}", styles['Normal'].clone('FooterStyle'))
    elements.append(footer)

    doc.build(elements)
    return response
